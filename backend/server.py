from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from starlette.responses import StreamingResponse
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import io
import csv
import json
import xmltodict
from openpyxl import load_workbook
import xlrd
import ftplib
import paramiko
import asyncio
from contextlib import asynccontextmanager
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from woocommerce import API as WooCommerceAPI
import requests

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'stockhub-secret-key-2024-secure-token')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Scheduler for automatic sync
scheduler = AsyncIOScheduler()

# Lifespan context manager for startup/shutdown
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    scheduler.add_job(sync_all_suppliers, 'interval', hours=12, id='sync_suppliers', replace_existing=True)
    scheduler.start()
    logger.info("Scheduler started - FTP sync every 12 hours")
    yield
    # Shutdown
    scheduler.shutdown()
    client.close()

app = FastAPI(title="StockHub SaaS API", lifespan=lifespan)
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== FTP SYNC FUNCTIONS ====================

def download_file_from_ftp_sync(supplier: dict) -> bytes:
    """Download file from FTP/SFTP server (synchronous)"""
    schema = supplier.get('ftp_schema', 'ftp').lower()
    host = supplier.get('ftp_host')
    port = supplier.get('ftp_port', 21)
    user = supplier.get('ftp_user', '')
    password = supplier.get('ftp_password', '')
    file_path = supplier.get('ftp_path', '')
    mode = supplier.get('ftp_mode', 'passive')
    
    if not host or not file_path:
        raise ValueError("FTP host and path are required")
    
    logger.info(f"Connecting to {schema.upper()}://{host}:{port}{file_path}")
    
    content = io.BytesIO()
    
    if schema == 'sftp':
        # SFTP connection using paramiko
        port = port or 22
        transport = paramiko.Transport((host, port))
        transport.connect(username=user, password=password)
        sftp = paramiko.SFTPClient.from_transport(transport)
        try:
            sftp.getfo(file_path, content)
            logger.info(f"SFTP download completed: {content.tell()} bytes")
        finally:
            sftp.close()
            transport.close()
    else:
        # FTP/FTPS connection
        port = port or 21
        if schema == 'ftps':
            ftp = ftplib.FTP_TLS()
        else:
            ftp = ftplib.FTP()
        
        try:
            ftp.connect(host, port, timeout=15)  # Reduced timeout to avoid Cloudflare 520
            ftp.login(user or 'anonymous', password or '')
            
            if schema == 'ftps':
                ftp.prot_p()  # Enable data channel encryption
            
            if mode == 'passive':
                ftp.set_pasv(True)
            else:
                ftp.set_pasv(False)
            
            logger.info(f"FTP connected, downloading {file_path}")
            ftp.retrbinary(f'RETR {file_path}', content.write)
            logger.info(f"FTP download completed: {content.tell()} bytes")
        finally:
            try:
                ftp.quit()
            except:
                pass
    
    content.seek(0)
    return content.read()

async def download_file_from_ftp(supplier: dict) -> bytes:
    """Async wrapper for FTP download"""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, download_file_from_ftp_sync, supplier)

def download_file_from_url_sync(url: str) -> bytes:
    """Download file from HTTP/HTTPS URL"""
    logger.info(f"Downloading from URL: {url}")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=60, stream=True)
        response.raise_for_status()
        
        content = response.content
        logger.info(f"URL download completed: {len(content)} bytes")
        return content
        
    except requests.exceptions.RequestException as e:
        logger.error(f"URL download failed: {e}")
        raise Exception(f"Error descargando desde URL: {str(e)}")

async def download_file_from_url(url: str) -> bytes:
    """Async wrapper for URL download"""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, download_file_from_url_sync, url)

def apply_column_mapping(raw_data: dict, column_mapping: dict) -> dict:
    """Apply column mapping to transform supplier data to system fields"""
    if not column_mapping:
        # No mapping, use auto-detection
        return normalize_product_data(raw_data)
    
    result = {}
    raw_lower = {str(k).lower().strip(): v for k, v in raw_data.items()}
    raw_original = {str(k).strip(): v for k, v in raw_data.items()}
    
    # System fields to map
    field_types = {
        'sku': 'string',
        'name': 'string',
        'description': 'string',
        'price': 'float',
        'price2': 'float',  # Secondary price
        'stock': 'int',
        'ean': 'string',
        'brand': 'string',
        'category': 'string',
        'subcategory': 'string',
        'subcategory2': 'string',
        'weight': 'float',
        'image_url': 'string',
        'image_url2': 'string',
        'image_url3': 'string',
        'short_description': 'string',
        'long_description': 'string'
    }
    
    for system_field, mapping in column_mapping.items():
        if not mapping:
            continue
            
        # mapping can be a string (single column) or list (multiple columns to concatenate)
        if isinstance(mapping, str):
            columns = [mapping]
        elif isinstance(mapping, list):
            columns = mapping
        else:
            continue
        
        values = []
        for col in columns:
            if not col:
                continue
            # Try to find the column (case insensitive)
            value = raw_original.get(col) or raw_lower.get(col.lower().strip())
            if value is not None and value != '':
                values.append(str(value))
        
        if values:
            combined_value = ' > '.join(values) if system_field.startswith('category') else ' '.join(values)
            
            # Convert to appropriate type
            field_type = field_types.get(system_field, 'string')
            try:
                if field_type == 'float':
                    combined_value = float(str(combined_value).replace(',', '.').replace('€', '').replace('$', '').strip())
                elif field_type == 'int':
                    combined_value = int(float(str(combined_value).replace(',', '.')))
            except:
                if field_type in ['float', 'int']:
                    combined_value = 0
            
            result[system_field] = combined_value
    
    # Build final product data
    product = {
        'sku': result.get('sku', ''),
        'name': result.get('name', ''),
        'description': result.get('description') or result.get('long_description') or result.get('short_description', ''),
        'price': result.get('price', 0),
        'stock': result.get('stock', 0),
        'category': result.get('category', ''),
        'brand': result.get('brand', ''),
        'ean': result.get('ean', ''),
        'weight': result.get('weight'),
        'image_url': result.get('image_url', '')
    }
    
    # Combine categories if multiple levels
    categories = [result.get('category', '')]
    if result.get('subcategory'):
        categories.append(result['subcategory'])
    if result.get('subcategory2'):
        categories.append(result['subcategory2'])
    product['category'] = ' > '.join([c for c in categories if c])
    
    return product

async def process_supplier_file(supplier: dict, content: bytes) -> dict:
    """Process downloaded file and update products"""
    file_format = supplier.get('file_format', 'csv').lower()
    separator = supplier.get('csv_separator', ';')
    enclosure = supplier.get('csv_enclosure', '"')
    header_row = supplier.get('csv_header_row', 1) or 1
    column_mapping = supplier.get('column_mapping')
    
    detected_columns = []
    
    try:
        if file_format == 'csv':
            # Handle CSV with custom settings
            try:
                decoded = content.decode('utf-8')
            except:
                decoded = content.decode('latin-1')
            
            lines = decoded.split('\n')
            if header_row > 1:
                lines = lines[header_row-1:]
            
            # Handle different separators
            if separator == '\\t':
                separator = '\t'
            
            reader = csv.DictReader(lines, delimiter=separator, quotechar=enclosure if enclosure else '"')
            raw_products = list(reader)
            
            # Get detected columns
            if raw_products:
                detected_columns = list(raw_products[0].keys())
                
        elif file_format in ['xlsx', 'xls']:
            if file_format == 'xlsx':
                wb = load_workbook(filename=io.BytesIO(content), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            else:
                wb = xlrd.open_workbook(file_contents=content)
                ws = wb.sheet_by_index(0)
                rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
            
            if not rows:
                return {"imported": 0, "updated": 0, "errors": 0}
            
            # Skip to header row
            if header_row > 1:
                rows = rows[header_row-1:]
            
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
            detected_columns = headers
            raw_products = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
            
        elif file_format == 'xml':
            raw_products = parse_xml_content(content)
            if raw_products:
                detected_columns = list(raw_products[0].keys())
        else:
            return {"imported": 0, "updated": 0, "errors": 0, "message": "Unsupported format"}
        
        # Save detected columns to supplier
        if detected_columns:
            await db.suppliers.update_one(
                {"id": supplier['id']},
                {"$set": {"detected_columns": detected_columns}}
            )
        
        # Process products
        now = datetime.now(timezone.utc).isoformat()
        imported = 0
        updated = 0
        errors = 0
        
        # If no column mapping and we have detected columns, warn the user
        needs_mapping = False
        if not column_mapping and detected_columns:
            # Check if auto-detection might work by looking at column names
            col_names_lower = [c.lower().strip() for c in detected_columns]
            has_sku = any(x in col_names_lower for x in ['sku', 'codigo', 'referencia', 'ref', 'reference', 'id'])
            has_name = any(x in col_names_lower for x in ['name', 'nombre', 'title', 'titulo', 'product', 'producto'])
            if not has_sku or not has_name:
                needs_mapping = True
        
        for raw in raw_products:
            try:
                # Apply column mapping or auto-detect
                if column_mapping:
                    normalized = apply_column_mapping(raw, column_mapping)
                else:
                    normalized = normalize_product_data(raw)
                    
                if not normalized.get('sku') or not normalized.get('name'):
                    errors += 1
                    continue
                
                existing = await db.products.find_one({
                    "sku": normalized['sku'], 
                    "supplier_id": supplier['id']
                })
                
                product_doc = {
                    "sku": normalized.get('sku'),
                    "name": normalized.get('name'),
                    "description": normalized.get('description'),
                    "price": normalized.get('price', 0),
                    "stock": normalized.get('stock', 0),
                    "category": normalized.get('category'),
                    "brand": normalized.get('brand'),
                    "ean": normalized.get('ean'),
                    "weight": normalized.get('weight'),
                    "image_url": normalized.get('image_url'),
                    "supplier_id": supplier['id'],
                    "supplier_name": supplier["name"],
                    "user_id": supplier["user_id"],
                    "updated_at": now
                }
                
                if existing:
                    # Track price changes
                    if existing.get('price') != product_doc['price'] and existing.get('price', 0) > 0:
                        await db.price_history.insert_one({
                            "id": str(uuid.uuid4()),
                            "product_id": existing["id"],
                            "product_name": product_doc["name"],
                            "old_price": existing.get('price', 0),
                            "new_price": product_doc['price'],
                            "change_percentage": ((product_doc['price'] - existing.get('price', 0)) / existing.get('price', 1)) * 100,
                            "user_id": supplier["user_id"],
                            "created_at": now
                        })
                    
                    # Track stock changes
                    if existing.get('stock', 0) > 0 and product_doc['stock'] == 0:
                        await db.notifications.insert_one({
                            "id": str(uuid.uuid4()),
                            "type": "stock_out",
                            "message": f"Producto '{product_doc['name']}' sin stock",
                            "product_id": existing["id"],
                            "product_name": product_doc["name"],
                            "user_id": supplier["user_id"],
                            "read": False,
                            "created_at": now
                        })
                    elif existing.get('stock', 0) > 5 and product_doc['stock'] <= 5 and product_doc['stock'] > 0:
                        await db.notifications.insert_one({
                            "id": str(uuid.uuid4()),
                            "type": "stock_low",
                            "message": f"Producto '{product_doc['name']}' con stock bajo ({product_doc['stock']} uds)",
                            "product_id": existing["id"],
                            "product_name": product_doc["name"],
                            "user_id": supplier["user_id"],
                            "read": False,
                            "created_at": now
                        })
                    
                    await db.products.update_one({"id": existing["id"]}, {"$set": product_doc})
                    updated += 1
                else:
                    product_doc["id"] = str(uuid.uuid4())
                    product_doc["created_at"] = now
                    await db.products.insert_one(product_doc)
                    imported += 1
            except Exception as e:
                logger.error(f"Error processing product: {e}")
                errors += 1
        
        result = {
            "imported": imported, 
            "updated": updated, 
            "errors": errors, 
            "detected_columns": detected_columns
        }
        
        # Add helpful message if mapping is needed
        if needs_mapping and errors > 0 and (imported + updated) == 0:
            result["message"] = "Se detectaron columnas pero no se pudieron importar productos. Configura el mapeo de columnas para asignar los campos correctamente."
            result["needs_mapping"] = True
        
        return result
    
    except Exception as e:
        logger.error(f"Error processing file: {e}")
        return {"imported": 0, "updated": 0, "errors": 1, "message": str(e), "detected_columns": []}

async def sync_supplier(supplier: dict) -> dict:
    """Sync a single supplier from FTP or URL"""
    connection_type = supplier.get('connection_type', 'ftp')
    
    # Check if connection is configured
    if connection_type == 'url':
        if not supplier.get('file_url'):
            return {"status": "skipped", "message": "URL no configurada"}
    else:  # FTP
        if not supplier.get('ftp_host') or not supplier.get('ftp_path'):
            return {"status": "skipped", "message": "FTP no configurado"}
    
    try:
        logger.info(f"Syncing supplier: {supplier['name']} (via {connection_type})")
        
        # Download file based on connection type
        if connection_type == 'url':
            content = await download_file_from_url(supplier['file_url'])
        else:
            content = await download_file_from_ftp(supplier)
        
        # Process file
        result = await process_supplier_file(supplier, content)
        
        # Update supplier stats
        now = datetime.now(timezone.utc).isoformat()
        product_count = await db.products.count_documents({"supplier_id": supplier['id']})
        await db.suppliers.update_one(
            {"id": supplier['id']},
            {"$set": {"product_count": product_count, "last_sync": now}}
        )
        
        # Create sync log notification
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "sync_complete",
            "message": f"Sincronización completada: {supplier['name']} - {result['imported']} nuevos, {result['updated']} actualizados",
            "product_id": None,
            "product_name": None,
            "user_id": supplier["user_id"],
            "read": False,
            "created_at": now
        })
        
        logger.info(f"Sync complete for {supplier['name']}: {result}")
        return {"status": "success", **result}
        
    except Exception as e:
        logger.error(f"Error syncing supplier {supplier['name']}: {e}")
        
        # Create error notification
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "sync_error",
            "message": f"Error en sincronización: {supplier['name']} - {str(e)[:100]}",
            "product_id": None,
            "product_name": None,
            "user_id": supplier["user_id"],
            "read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"status": "error", "message": str(e)}

async def sync_all_suppliers():
    """Sync all suppliers with FTP configured - runs every 12 hours"""
    logger.info("Starting scheduled sync for all suppliers...")
    
    # Find suppliers with FTP configured
    ftp_suppliers = await db.suppliers.find({
        "connection_type": {"$ne": "url"},
        "ftp_host": {"$ne": None, "$ne": ""},
        "ftp_path": {"$ne": None, "$ne": ""}
    }).to_list(1000)
    
    # Find suppliers with URL configured
    url_suppliers = await db.suppliers.find({
        "connection_type": "url",
        "file_url": {"$ne": None, "$ne": ""}
    }).to_list(1000)
    
    all_suppliers = ftp_suppliers + url_suppliers
    logger.info(f"Found {len(all_suppliers)} suppliers to sync ({len(ftp_suppliers)} FTP, {len(url_suppliers)} URL)")
    
    for supplier in all_suppliers:
        await sync_supplier(supplier)
        await asyncio.sleep(2)  # Small delay between syncs
    
    logger.info("Scheduled sync completed")

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    company: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    company: Optional[str] = None
    created_at: str

class SupplierCreate(BaseModel):
    name: str
    description: Optional[str] = None
    # Tipo de conexión
    connection_type: Optional[str] = "ftp"  # ftp, url
    # URL directa del archivo
    file_url: Optional[str] = None
    # Conexión FTP
    ftp_schema: Optional[str] = "ftp"  # ftp, sftp, ftps
    ftp_host: Optional[str] = None
    ftp_user: Optional[str] = None
    ftp_password: Optional[str] = None
    ftp_port: Optional[int] = 21
    ftp_path: Optional[str] = None
    ftp_mode: Optional[str] = "passive"  # passive, active
    # Configuración CSV
    file_format: Optional[str] = "csv"  # csv, xlsx, xls, xml
    csv_separator: Optional[str] = ";"
    csv_enclosure: Optional[str] = '"'
    csv_line_break: Optional[str] = "\\n"
    csv_header_row: Optional[int] = 1
    # Mapeo de columnas - campo del sistema: columna(s) del proveedor
    column_mapping: Optional[Dict[str, Any]] = None

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    # Tipo de conexión
    connection_type: Optional[str] = None
    # URL directa del archivo
    file_url: Optional[str] = None
    # Conexión FTP
    ftp_schema: Optional[str] = None
    ftp_host: Optional[str] = None
    ftp_user: Optional[str] = None
    ftp_password: Optional[str] = None
    ftp_port: Optional[int] = None
    ftp_path: Optional[str] = None
    ftp_mode: Optional[str] = None
    # Configuración CSV
    file_format: Optional[str] = None
    csv_separator: Optional[str] = None
    csv_enclosure: Optional[str] = None
    csv_line_break: Optional[str] = None
    csv_header_row: Optional[int] = None
    # Mapeo de columnas
    column_mapping: Optional[Dict[str, Any]] = None

class SupplierResponse(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    # Tipo de conexión
    connection_type: Optional[str] = "ftp"
    # URL directa
    file_url: Optional[str] = None
    # Conexión FTP
    ftp_schema: Optional[str] = None
    ftp_host: Optional[str] = None
    ftp_user: Optional[str] = None
    ftp_port: Optional[int] = None
    ftp_path: Optional[str] = None
    ftp_mode: Optional[str] = None
    # Configuración CSV
    file_format: Optional[str] = None
    csv_separator: Optional[str] = None
    csv_enclosure: Optional[str] = None
    csv_line_break: Optional[str] = None
    csv_header_row: Optional[int] = None
    column_mapping: Optional[Dict[str, Any]] = None
    # Detected columns from last import
    detected_columns: Optional[List[str]] = None
    # Stats
    product_count: int = 0
    last_sync: Optional[str] = None
    created_at: str

class ProductBase(BaseModel):
    sku: str
    name: str
    description: Optional[str] = None
    price: float
    stock: int
    category: Optional[str] = None
    brand: Optional[str] = None
    ean: Optional[str] = None
    weight: Optional[float] = None
    image_url: Optional[str] = None
    attributes: Optional[Dict[str, Any]] = None

class ProductResponse(ProductBase):
    id: str
    supplier_id: str
    supplier_name: str
    created_at: str
    updated_at: str

class CatalogItemCreate(BaseModel):
    product_id: str
    custom_price: Optional[float] = None
    custom_name: Optional[str] = None
    active: bool = True

class CatalogItemResponse(BaseModel):
    id: str
    product_id: str
    product: ProductResponse
    custom_price: Optional[float] = None
    custom_name: Optional[str] = None
    final_price: float
    active: bool
    created_at: str

class MarginRuleCreate(BaseModel):
    name: str
    rule_type: str  # percentage, fixed, tiered
    value: float  # percentage or fixed amount
    apply_to: str  # all, category, supplier, product
    apply_to_value: Optional[str] = None  # category name, supplier id, product id
    min_price: Optional[float] = None
    max_price: Optional[float] = None
    priority: int = 0

class MarginRuleResponse(MarginRuleCreate):
    id: str
    user_id: str
    created_at: str

class NotificationResponse(BaseModel):
    id: str
    type: str  # stock_low, stock_out, price_change, sync_complete, sync_error
    message: str
    product_id: Optional[str] = None
    product_name: Optional[str] = None
    read: bool
    created_at: str

class PriceHistoryResponse(BaseModel):
    id: str
    product_id: str
    product_name: str
    old_price: float
    new_price: float
    change_percentage: float
    created_at: str

class ExportRequest(BaseModel):
    platform: str  # prestashop, woocommerce, shopify
    catalog_ids: Optional[List[str]] = None

# ==================== CATALOG MODELS ====================

class CatalogCreate(BaseModel):
    name: str = Field(..., description="Nombre del catálogo")
    description: Optional[str] = None
    is_default: bool = False

class CatalogUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    is_default: Optional[bool] = None

class CatalogResponse(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    is_default: bool = False
    product_count: int = 0
    margin_rules_count: int = 0
    created_at: str

class CatalogProductAdd(BaseModel):
    product_ids: List[str]
    custom_prices: Optional[Dict[str, float]] = None  # product_id: custom_price

class CatalogMarginRuleCreate(BaseModel):
    catalog_id: str
    name: str
    rule_type: str = "percentage"  # percentage, fixed
    value: float
    apply_to: str = "all"  # all, category, supplier, brand
    apply_to_value: Optional[str] = None
    priority: int = 0

class CatalogMarginRuleResponse(BaseModel):
    id: str
    catalog_id: str
    name: str
    rule_type: str
    value: float
    apply_to: str
    apply_to_value: Optional[str] = None
    priority: int
    created_at: str

# ==================== WOOCOMMERCE INTEGRATION ====================

class WooCommerceConfig(BaseModel):
    store_url: str = Field(..., description="URL de la tienda WooCommerce (ej: https://mitienda.com)")
    consumer_key: str = Field(..., description="Consumer Key de la API REST")
    consumer_secret: str = Field(..., description="Consumer Secret de la API REST")
    name: Optional[str] = "Mi Tienda WooCommerce"

class WooCommerceConfigUpdate(BaseModel):
    store_url: Optional[str] = None
    consumer_key: Optional[str] = None
    consumer_secret: Optional[str] = None
    name: Optional[str] = None

class WooCommerceConfigResponse(BaseModel):
    id: str
    name: str
    store_url: str
    consumer_key_masked: str  # Solo mostrar últimos 4 caracteres
    is_connected: bool = False
    last_sync: Optional[str] = None
    products_synced: int = 0
    created_at: str

class WooCommerceExportRequest(BaseModel):
    config_id: str
    catalog_id: Optional[str] = None  # ID del catálogo a exportar. Si es None, exportar catálogo por defecto
    update_existing: bool = True  # Actualizar productos existentes por SKU

class WooCommerceExportResult(BaseModel):
    status: str
    created: int = 0
    updated: int = 0
    failed: int = 0
    errors: List[str] = []

class DashboardStats(BaseModel):
    total_suppliers: int
    total_products: int
    total_catalog_items: int
    low_stock_count: int
    out_of_stock_count: int
    unread_notifications: int
    recent_price_changes: int

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str) -> str:
    payload = {
        "user_id": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=dict)
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": user.email,
        "password": hash_password(user.password),
        "name": user.name,
        "company": user.company,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    token = create_token(user_id)
    return {"token": token, "user": {"id": user_id, "email": user.email, "name": user.name, "company": user.company}}

@api_router.post("/auth/login", response_model=dict)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    token = create_token(user["id"])
    return {"token": token, "user": {"id": user["id"], "email": user["email"], "name": user["name"], "company": user.get("company")}}

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(**user)

# ==================== SUPPLIERS ENDPOINTS ====================

@api_router.post("/suppliers", response_model=SupplierResponse)
async def create_supplier(supplier: SupplierCreate, user: dict = Depends(get_current_user)):
    supplier_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    supplier_doc = {
        "id": supplier_id,
        "user_id": user["id"],
        "name": supplier.name,
        "description": supplier.description,
        # Tipo de conexión
        "connection_type": supplier.connection_type or "ftp",
        # URL directa
        "file_url": supplier.file_url,
        # Conexión FTP
        "ftp_schema": supplier.ftp_schema,
        "ftp_host": supplier.ftp_host,
        "ftp_user": supplier.ftp_user,
        "ftp_password": supplier.ftp_password,
        "ftp_port": supplier.ftp_port,
        "ftp_path": supplier.ftp_path,
        "ftp_mode": supplier.ftp_mode,
        # Configuración CSV
        "file_format": supplier.file_format,
        "csv_separator": supplier.csv_separator,
        "csv_enclosure": supplier.csv_enclosure,
        "csv_line_break": supplier.csv_line_break,
        "csv_header_row": supplier.csv_header_row,
        "column_mapping": supplier.column_mapping,
        # Stats
        "product_count": 0,
        "last_sync": None,
        "created_at": now
    }
    await db.suppliers.insert_one(supplier_doc)
    supplier_doc.pop("ftp_password", None)
    supplier_doc.pop("user_id", None)
    supplier_doc.pop("_id", None)
    return SupplierResponse(**supplier_doc)

@api_router.get("/suppliers", response_model=List[SupplierResponse])
async def get_suppliers(user: dict = Depends(get_current_user)):
    suppliers = await db.suppliers.find({"user_id": user["id"]}, {"_id": 0, "ftp_password": 0, "user_id": 0}).to_list(1000)
    return [SupplierResponse(**s) for s in suppliers]

@api_router.get("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def get_supplier(supplier_id: str, user: dict = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id, "user_id": user["id"]}, {"_id": 0, "ftp_password": 0, "user_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return SupplierResponse(**supplier)

@api_router.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(supplier_id: str, supplier: SupplierUpdate, user: dict = Depends(get_current_user)):
    existing = await db.suppliers.find_one({"id": supplier_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    update_data = {k: v for k, v in supplier.model_dump().items() if v is not None}
    if update_data:
        await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    
    updated = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0, "ftp_password": 0, "user_id": 0})
    return SupplierResponse(**updated)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, user: dict = Depends(get_current_user)):
    result = await db.suppliers.delete_one({"id": supplier_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    await db.products.delete_many({"supplier_id": supplier_id})
    return {"message": "Proveedor eliminado"}

@api_router.post("/suppliers/{supplier_id}/sync")
async def sync_supplier_manual(supplier_id: str, user: dict = Depends(get_current_user)):
    """Manually trigger sync for a supplier (FTP or URL)"""
    supplier = await db.suppliers.find_one({"id": supplier_id, "user_id": user["id"]})
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    connection_type = supplier.get('connection_type', 'ftp')
    
    # Validate connection configuration
    if connection_type == 'url':
        if not supplier.get('file_url'):
            raise HTTPException(status_code=400, detail="URL del archivo no configurada.")
    else:  # FTP
        if not supplier.get('ftp_host') or not supplier.get('ftp_path'):
            raise HTTPException(status_code=400, detail="Configuración FTP incompleta. Configure Host y Ruta del archivo.")
    
    try:
        result = await sync_supplier(supplier)
        
        if result.get('status') == 'error':
            return {"status": "error", "message": result.get('message', 'Error en sincronización')}
        
        return result
    except Exception as e:
        logger.error(f"Exception in sync_supplier_manual: {e}")
        return {"status": "error", "message": str(e)}

@api_router.get("/suppliers/{supplier_id}/sync-status")
async def get_sync_status(supplier_id: str, user: dict = Depends(get_current_user)):
    """Get sync status for a supplier"""
    supplier = await db.suppliers.find_one({"id": supplier_id, "user_id": user["id"]}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    # Get next scheduled sync time
    job = scheduler.get_job('sync_suppliers')
    next_run = job.next_run_time.isoformat() if job and job.next_run_time else None
    
    return {
        "last_sync": supplier.get('last_sync'),
        "next_scheduled_sync": next_run,
        "ftp_configured": bool(supplier.get('ftp_host') and supplier.get('ftp_path')),
        "product_count": supplier.get('product_count', 0)
    }

# ==================== FILE PARSING HELPERS ====================

def parse_csv_content(content: bytes) -> List[dict]:
    try:
        decoded = content.decode('utf-8')
    except:
        decoded = content.decode('latin-1')
    reader = csv.DictReader(io.StringIO(decoded))
    return list(reader)

def parse_xlsx_content(content: bytes) -> List[dict]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).lower().strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(row)]

def parse_xls_content(content: bytes) -> List[dict]:
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, c)).lower().strip() for c in range(ws.ncols)]
    result = []
    for r in range(1, ws.nrows):
        row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
        result.append(row)
    return result

def parse_xml_content(content: bytes) -> List[dict]:
    try:
        decoded = content.decode('utf-8')
    except:
        decoded = content.decode('latin-1')
    data = xmltodict.parse(decoded)
    # Try to find the products list
    for key in ['products', 'items', 'catalog', 'data', 'root']:
        if key in data:
            items = data[key]
            if isinstance(items, dict):
                for subkey in items:
                    if isinstance(items[subkey], list):
                        return items[subkey]
            elif isinstance(items, list):
                return items
    return []

def normalize_product_data(raw: dict) -> dict:
    """Normalize different field names to our standard format"""
    mapping = {
        'sku': ['sku', 'codigo', 'code', 'ref', 'referencia', 'reference', 'id', 'product_id'],
        'name': ['name', 'nombre', 'title', 'titulo', 'product_name', 'description'],
        'price': ['price', 'precio', 'pvp', 'cost', 'coste', 'unit_price'],
        'stock': ['stock', 'quantity', 'cantidad', 'qty', 'inventory', 'disponible'],
        'category': ['category', 'categoria', 'cat', 'type', 'tipo'],
        'brand': ['brand', 'marca', 'manufacturer', 'fabricante'],
        'ean': ['ean', 'ean13', 'barcode', 'upc', 'codigo_barras'],
        'weight': ['weight', 'peso', 'kg'],
        'image_url': ['image', 'imagen', 'image_url', 'photo', 'foto', 'picture'],
        'description': ['description', 'descripcion', 'desc', 'details', 'detalles']
    }
    
    result = {}
    raw_lower = {str(k).lower().strip(): v for k, v in raw.items()}
    
    for field, aliases in mapping.items():
        for alias in aliases:
            if alias in raw_lower and raw_lower[alias]:
                value = raw_lower[alias]
                if field in ['price', 'weight']:
                    try:
                        value = float(str(value).replace(',', '.').replace('€', '').strip())
                    except:
                        value = 0.0
                elif field == 'stock':
                    try:
                        value = int(float(str(value).replace(',', '.')))
                    except:
                        value = 0
                result[field] = value
                break
    
    return result

# ==================== PRODUCTS ENDPOINTS ====================

@api_router.post("/products/import/{supplier_id}")
async def import_products(supplier_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id, "user_id": user["id"]})
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    content = await file.read()
    filename = file.filename.lower()
    
    try:
        if filename.endswith('.csv'):
            raw_products = parse_csv_content(content)
        elif filename.endswith('.xlsx'):
            raw_products = parse_xlsx_content(content)
        elif filename.endswith('.xls'):
            raw_products = parse_xls_content(content)
        elif filename.endswith('.xml'):
            raw_products = parse_xml_content(content)
        else:
            raise HTTPException(status_code=400, detail="Formato de archivo no soportado. Use CSV, XLSX, XLS o XML")
    except Exception as e:
        logger.error(f"Error parsing file: {e}")
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(e)}")
    
    now = datetime.now(timezone.utc).isoformat()
    imported = 0
    updated = 0
    
    for raw in raw_products:
        normalized = normalize_product_data(raw)
        if not normalized.get('sku') or not normalized.get('name'):
            continue
        
        existing = await db.products.find_one({"sku": normalized['sku'], "supplier_id": supplier_id})
        
        product_doc = {
            "sku": normalized.get('sku'),
            "name": normalized.get('name'),
            "description": normalized.get('description'),
            "price": normalized.get('price', 0),
            "stock": normalized.get('stock', 0),
            "category": normalized.get('category'),
            "brand": normalized.get('brand'),
            "ean": normalized.get('ean'),
            "weight": normalized.get('weight'),
            "image_url": normalized.get('image_url'),
            "supplier_id": supplier_id,
            "supplier_name": supplier["name"],
            "user_id": user["id"],
            "updated_at": now
        }
        
        if existing:
            # Track price changes
            if existing.get('price') != product_doc['price']:
                await db.price_history.insert_one({
                    "id": str(uuid.uuid4()),
                    "product_id": existing["id"],
                    "product_name": product_doc["name"],
                    "old_price": existing.get('price', 0),
                    "new_price": product_doc['price'],
                    "change_percentage": ((product_doc['price'] - existing.get('price', 0)) / existing.get('price', 1)) * 100 if existing.get('price', 0) > 0 else 0,
                    "user_id": user["id"],
                    "created_at": now
                })
            
            # Track stock changes
            if existing.get('stock', 0) > 0 and product_doc['stock'] == 0:
                await db.notifications.insert_one({
                    "id": str(uuid.uuid4()),
                    "type": "stock_out",
                    "message": f"Producto '{product_doc['name']}' sin stock",
                    "product_id": existing["id"],
                    "product_name": product_doc["name"],
                    "user_id": user["id"],
                    "read": False,
                    "created_at": now
                })
            elif existing.get('stock', 0) > 5 and product_doc['stock'] <= 5 and product_doc['stock'] > 0:
                await db.notifications.insert_one({
                    "id": str(uuid.uuid4()),
                    "type": "stock_low",
                    "message": f"Producto '{product_doc['name']}' con stock bajo ({product_doc['stock']} unidades)",
                    "product_id": existing["id"],
                    "product_name": product_doc["name"],
                    "user_id": user["id"],
                    "read": False,
                    "created_at": now
                })
            
            await db.products.update_one({"id": existing["id"]}, {"$set": product_doc})
            updated += 1
        else:
            product_doc["id"] = str(uuid.uuid4())
            product_doc["created_at"] = now
            await db.products.insert_one(product_doc)
            imported += 1
    
    # Update supplier stats
    product_count = await db.products.count_documents({"supplier_id": supplier_id})
    await db.suppliers.update_one({"id": supplier_id}, {"$set": {"product_count": product_count, "last_sync": now}})
    
    return {"imported": imported, "updated": updated, "total": imported + updated}

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(
    supplier_id: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    min_stock: Optional[int] = None,
    max_stock: Optional[int] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    skip: int = 0,
    limit: int = 100,
    user: dict = Depends(get_current_user)
):
    query = {"user_id": user["id"]}
    if supplier_id:
        query["supplier_id"] = supplier_id
    if category:
        query["category"] = category
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"ean": {"$regex": search, "$options": "i"}}
        ]
    if min_stock is not None:
        query["stock"] = {"$gte": min_stock}
    if max_stock is not None:
        query.setdefault("stock", {})["$lte"] = max_stock
    if min_price is not None:
        query["price"] = {"$gte": min_price}
    if max_price is not None:
        query.setdefault("price", {})["$lte"] = max_price
    
    products = await db.products.find(query, {"_id": 0, "user_id": 0}).skip(skip).limit(limit).to_list(limit)
    return [ProductResponse(**p) for p in products]

@api_router.get("/products/categories")
async def get_categories(user: dict = Depends(get_current_user)):
    categories = await db.products.distinct("category", {"user_id": user["id"], "category": {"$ne": None}})
    return categories

@api_router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str, user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id, "user_id": user["id"]}, {"_id": 0, "user_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return ProductResponse(**product)

# ==================== CATALOGS MANAGEMENT ====================

@api_router.post("/catalogs", response_model=CatalogResponse)
async def create_catalog(catalog: CatalogCreate, user: dict = Depends(get_current_user)):
    """Create a new catalog"""
    catalog_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # If this is the default catalog, unset other defaults
    if catalog.is_default:
        await db.catalogs.update_many(
            {"user_id": user["id"]},
            {"$set": {"is_default": False}}
        )
    
    # Check if user has no catalogs - make first one default
    existing_count = await db.catalogs.count_documents({"user_id": user["id"]})
    is_default = catalog.is_default or existing_count == 0
    
    catalog_doc = {
        "id": catalog_id,
        "user_id": user["id"],
        "name": catalog.name,
        "description": catalog.description,
        "is_default": is_default,
        "created_at": now
    }
    await db.catalogs.insert_one(catalog_doc)
    
    return CatalogResponse(
        id=catalog_id,
        name=catalog.name,
        description=catalog.description,
        is_default=is_default,
        product_count=0,
        margin_rules_count=0,
        created_at=now
    )

@api_router.get("/catalogs", response_model=List[CatalogResponse])
async def get_catalogs(user: dict = Depends(get_current_user)):
    """Get all catalogs for user"""
    catalogs = await db.catalogs.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    
    result = []
    for cat in catalogs:
        product_count = await db.catalog_items.count_documents({"catalog_id": cat["id"]})
        rules_count = await db.catalog_margin_rules.count_documents({"catalog_id": cat["id"]})
        result.append(CatalogResponse(
            id=cat["id"],
            name=cat["name"],
            description=cat.get("description"),
            is_default=cat.get("is_default", False),
            product_count=product_count,
            margin_rules_count=rules_count,
            created_at=cat["created_at"]
        ))
    
    return result

@api_router.get("/catalogs/{catalog_id}", response_model=CatalogResponse)
async def get_catalog_by_id(catalog_id: str, user: dict = Depends(get_current_user)):
    """Get a specific catalog"""
    catalog = await db.catalogs.find_one({"id": catalog_id, "user_id": user["id"]}, {"_id": 0})
    if not catalog:
        raise HTTPException(status_code=404, detail="Catálogo no encontrado")
    
    product_count = await db.catalog_items.count_documents({"catalog_id": catalog_id})
    rules_count = await db.catalog_margin_rules.count_documents({"catalog_id": catalog_id})
    
    return CatalogResponse(
        id=catalog["id"],
        name=catalog["name"],
        description=catalog.get("description"),
        is_default=catalog.get("is_default", False),
        product_count=product_count,
        margin_rules_count=rules_count,
        created_at=catalog["created_at"]
    )

@api_router.put("/catalogs/{catalog_id}", response_model=CatalogResponse)
async def update_catalog(catalog_id: str, update: CatalogUpdate, user: dict = Depends(get_current_user)):
    """Update a catalog"""
    existing = await db.catalogs.find_one({"id": catalog_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Catálogo no encontrado")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    # If setting as default, unset other defaults
    if update_data.get("is_default"):
        await db.catalogs.update_many(
            {"user_id": user["id"], "id": {"$ne": catalog_id}},
            {"$set": {"is_default": False}}
        )
    
    if update_data:
        await db.catalogs.update_one({"id": catalog_id}, {"$set": update_data})
    
    updated = await db.catalogs.find_one({"id": catalog_id}, {"_id": 0})
    product_count = await db.catalog_items.count_documents({"catalog_id": catalog_id})
    rules_count = await db.catalog_margin_rules.count_documents({"catalog_id": catalog_id})
    
    return CatalogResponse(
        id=updated["id"],
        name=updated["name"],
        description=updated.get("description"),
        is_default=updated.get("is_default", False),
        product_count=product_count,
        margin_rules_count=rules_count,
        created_at=updated["created_at"]
    )

@api_router.delete("/catalogs/{catalog_id}")
async def delete_catalog(catalog_id: str, user: dict = Depends(get_current_user)):
    """Delete a catalog and its items"""
    existing = await db.catalogs.find_one({"id": catalog_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Catálogo no encontrado")
    
    # Delete catalog items
    await db.catalog_items.delete_many({"catalog_id": catalog_id})
    # Delete margin rules
    await db.catalog_margin_rules.delete_many({"catalog_id": catalog_id})
    # Delete catalog
    await db.catalogs.delete_one({"id": catalog_id})
    
    return {"message": "Catálogo eliminado"}

# ==================== CATALOG ITEMS (Products in Catalog) ====================

@api_router.post("/catalogs/{catalog_id}/products")
async def add_products_to_catalog(catalog_id: str, data: CatalogProductAdd, user: dict = Depends(get_current_user)):
    """Add products to a catalog"""
    catalog = await db.catalogs.find_one({"id": catalog_id, "user_id": user["id"]})
    if not catalog:
        raise HTTPException(status_code=404, detail="Catálogo no encontrado")
    
    added = 0
    for product_id in data.product_ids:
        # Check product exists
        product = await db.products.find_one({"id": product_id, "user_id": user["id"]})
        if not product:
            continue
        
        # Check not already in catalog
        existing = await db.catalog_items.find_one({"catalog_id": catalog_id, "product_id": product_id})
        if existing:
            continue
        
        item_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()
        custom_price = data.custom_prices.get(product_id) if data.custom_prices else None
        
        await db.catalog_items.insert_one({
            "id": item_id,
            "catalog_id": catalog_id,
            "product_id": product_id,
            "user_id": user["id"],
            "custom_price": custom_price,
            "custom_name": None,
            "active": True,
            "created_at": now
        })
        added += 1
    
    return {"added": added, "message": f"{added} productos añadidos al catálogo"}

@api_router.get("/catalogs/{catalog_id}/products")
async def get_catalog_products(
    catalog_id: str,
    active_only: bool = False,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    user: dict = Depends(get_current_user)
):
    """Get products in a catalog with calculated prices"""
    catalog = await db.catalogs.find_one({"id": catalog_id, "user_id": user["id"]})
    if not catalog:
        raise HTTPException(status_code=404, detail="Catálogo no encontrado")
    
    query = {"catalog_id": catalog_id}
    if active_only:
        query["active"] = True
    
    items = await db.catalog_items.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    # Get catalog-specific margin rules
    margin_rules = await db.catalog_margin_rules.find(
        {"catalog_id": catalog_id}, {"_id": 0}
    ).sort("priority", -1).to_list(100)
    
    result = []
    for item in items:
        product = await db.products.find_one({"id": item["product_id"]}, {"_id": 0, "user_id": 0})
        if not product:
            continue
        
        if search:
            search_lower = search.lower()
            if search_lower not in product.get("name", "").lower() and search_lower not in product.get("sku", "").lower():
                continue
        
        # Calculate final price using catalog-specific rules
        base_price = item.get("custom_price") or product.get("price", 0)
        final_price = calculate_final_price(base_price, product, margin_rules)
        
        result.append({
            "id": item["id"],
            "catalog_id": catalog_id,
            "product_id": item["product_id"],
            "product": product,
            "custom_price": item.get("custom_price"),
            "custom_name": item.get("custom_name"),
            "active": item.get("active", True),
            "final_price": final_price,
            "created_at": item["created_at"]
        })
    
    return result

@api_router.delete("/catalogs/{catalog_id}/products/{item_id}")
async def remove_product_from_catalog(catalog_id: str, item_id: str, user: dict = Depends(get_current_user)):
    """Remove a product from a catalog"""
    result = await db.catalog_items.delete_one({"id": item_id, "catalog_id": catalog_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Producto no encontrado en el catálogo")
    return {"message": "Producto eliminado del catálogo"}

# ==================== CATALOG MARGIN RULES ====================

@api_router.post("/catalogs/{catalog_id}/margin-rules", response_model=CatalogMarginRuleResponse)
async def create_catalog_margin_rule(catalog_id: str, rule: CatalogMarginRuleCreate, user: dict = Depends(get_current_user)):
    """Create a margin rule for a specific catalog"""
    catalog = await db.catalogs.find_one({"id": catalog_id, "user_id": user["id"]})
    if not catalog:
        raise HTTPException(status_code=404, detail="Catálogo no encontrado")
    
    rule_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    rule_doc = {
        "id": rule_id,
        "catalog_id": catalog_id,
        "user_id": user["id"],
        "name": rule.name,
        "rule_type": rule.rule_type,
        "value": rule.value,
        "apply_to": rule.apply_to,
        "apply_to_value": rule.apply_to_value,
        "priority": rule.priority,
        "created_at": now
    }
    await db.catalog_margin_rules.insert_one(rule_doc)
    
    return CatalogMarginRuleResponse(**{k: v for k, v in rule_doc.items() if k != "user_id"})

@api_router.get("/catalogs/{catalog_id}/margin-rules", response_model=List[CatalogMarginRuleResponse])
async def get_catalog_margin_rules(catalog_id: str, user: dict = Depends(get_current_user)):
    """Get margin rules for a specific catalog"""
    catalog = await db.catalogs.find_one({"id": catalog_id, "user_id": user["id"]})
    if not catalog:
        raise HTTPException(status_code=404, detail="Catálogo no encontrado")
    
    rules = await db.catalog_margin_rules.find(
        {"catalog_id": catalog_id}, {"_id": 0, "user_id": 0}
    ).sort("priority", -1).to_list(100)
    
    return [CatalogMarginRuleResponse(**r) for r in rules]

@api_router.delete("/catalogs/{catalog_id}/margin-rules/{rule_id}")
async def delete_catalog_margin_rule(catalog_id: str, rule_id: str, user: dict = Depends(get_current_user)):
    """Delete a margin rule from a catalog"""
    result = await db.catalog_margin_rules.delete_one({"id": rule_id, "catalog_id": catalog_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Regla no encontrada")
    return {"message": "Regla eliminada"}

# ==================== LEGACY CATALOG ENDPOINTS (for backward compatibility) ====================

@api_router.post("/catalog")
async def add_to_catalog(item: CatalogItemCreate, user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": item.product_id, "user_id": user["id"]})
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    existing = await db.catalog.find_one({"product_id": item.product_id, "user_id": user["id"]})
    if existing:
        raise HTTPException(status_code=400, detail="Producto ya está en el catálogo")
    
    catalog_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    catalog_doc = {
        "id": catalog_id,
        "product_id": item.product_id,
        "user_id": user["id"],
        "custom_price": item.custom_price,
        "custom_name": item.custom_name,
        "active": item.active,
        "created_at": now
    }
    await db.catalog.insert_one(catalog_doc)
    return {"id": catalog_id, "message": "Producto añadido al catálogo"}

@api_router.get("/catalog")
async def get_catalog(
    active_only: bool = False,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    user: dict = Depends(get_current_user)
):
    query = {"user_id": user["id"]}
    if active_only:
        query["active"] = True
    
    catalog_items = await db.catalog.find(query, {"_id": 0, "user_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    # Get margin rules for price calculation
    margin_rules = await db.margin_rules.find({"user_id": user["id"]}, {"_id": 0}).sort("priority", -1).to_list(100)
    
    result = []
    for item in catalog_items:
        product = await db.products.find_one({"id": item["product_id"]}, {"_id": 0, "user_id": 0})
        if not product:
            continue
        
        if search and search.lower() not in product.get("name", "").lower() and search.lower() not in product.get("sku", "").lower():
            continue
        
        # Calculate final price
        base_price = item.get("custom_price") or product.get("price", 0)
        final_price = calculate_final_price(base_price, product, margin_rules)
        
        result.append({
            "id": item["id"],
            "product_id": item["product_id"],
            "product": ProductResponse(**product),
            "custom_price": item.get("custom_price"),
            "custom_name": item.get("custom_name"),
            "final_price": round(final_price, 2),
            "active": item.get("active", True),
            "created_at": item.get("created_at")
        })
    
    return result

def calculate_final_price(base_price: float, product: dict, rules: List[dict]) -> float:
    """Calculate final price applying margin rules"""
    final_price = base_price
    
    for rule in rules:
        applies = False
        if rule["apply_to"] == "all":
            applies = True
        elif rule["apply_to"] == "category" and product.get("category") == rule.get("apply_to_value"):
            applies = True
        elif rule["apply_to"] == "supplier" and product.get("supplier_id") == rule.get("apply_to_value"):
            applies = True
        elif rule["apply_to"] == "product" and product.get("id") == rule.get("apply_to_value"):
            applies = True
        
        if applies:
            if rule.get("min_price") and base_price < rule["min_price"]:
                continue
            if rule.get("max_price") and base_price > rule["max_price"]:
                continue
            
            if rule["rule_type"] == "percentage":
                final_price = base_price * (1 + rule["value"] / 100)
            elif rule["rule_type"] == "fixed":
                final_price = base_price + rule["value"]
            break  # Apply first matching rule
    
    return final_price

@api_router.put("/catalog/{catalog_id}")
async def update_catalog_item(catalog_id: str, item: CatalogItemCreate, user: dict = Depends(get_current_user)):
    existing = await db.catalog.find_one({"id": catalog_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    
    update_data = {
        "custom_price": item.custom_price,
        "custom_name": item.custom_name,
        "active": item.active
    }
    await db.catalog.update_one({"id": catalog_id}, {"$set": update_data})
    return {"message": "Item actualizado"}

@api_router.delete("/catalog/{catalog_id}")
async def remove_from_catalog(catalog_id: str, user: dict = Depends(get_current_user)):
    result = await db.catalog.delete_one({"id": catalog_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    return {"message": "Item eliminado del catálogo"}

# ==================== MARGIN RULES ENDPOINTS ====================

@api_router.post("/margin-rules", response_model=MarginRuleResponse)
async def create_margin_rule(rule: MarginRuleCreate, user: dict = Depends(get_current_user)):
    rule_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    rule_doc = {
        "id": rule_id,
        "user_id": user["id"],
        **rule.model_dump(),
        "created_at": now
    }
    await db.margin_rules.insert_one(rule_doc)
    rule_doc.pop("_id", None)
    return MarginRuleResponse(**rule_doc)

@api_router.get("/margin-rules", response_model=List[MarginRuleResponse])
async def get_margin_rules(user: dict = Depends(get_current_user)):
    rules = await db.margin_rules.find({"user_id": user["id"]}, {"_id": 0}).sort("priority", -1).to_list(100)
    return [MarginRuleResponse(**r) for r in rules]

@api_router.put("/margin-rules/{rule_id}", response_model=MarginRuleResponse)
async def update_margin_rule(rule_id: str, rule: MarginRuleCreate, user: dict = Depends(get_current_user)):
    existing = await db.margin_rules.find_one({"id": rule_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Regla no encontrada")
    
    await db.margin_rules.update_one({"id": rule_id}, {"$set": rule.model_dump()})
    updated = await db.margin_rules.find_one({"id": rule_id}, {"_id": 0})
    return MarginRuleResponse(**updated)

@api_router.delete("/margin-rules/{rule_id}")
async def delete_margin_rule(rule_id: str, user: dict = Depends(get_current_user)):
    result = await db.margin_rules.delete_one({"id": rule_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Regla no encontrada")
    return {"message": "Regla eliminada"}

# ==================== EXPORT ENDPOINTS ====================

@api_router.post("/export")
async def export_catalog(request: ExportRequest, user: dict = Depends(get_current_user)):
    query = {"user_id": user["id"], "active": True}
    if request.catalog_ids:
        query["id"] = {"$in": request.catalog_ids}
    
    catalog_items = await db.catalog.find(query, {"_id": 0}).to_list(10000)
    margin_rules = await db.margin_rules.find({"user_id": user["id"]}, {"_id": 0}).sort("priority", -1).to_list(100)
    
    rows = []
    for item in catalog_items:
        product = await db.products.find_one({"id": item["product_id"]}, {"_id": 0, "user_id": 0})
        if not product:
            continue
        
        base_price = item.get("custom_price") or product.get("price", 0)
        final_price = calculate_final_price(base_price, product, margin_rules)
        
        name = item.get("custom_name") or product.get("name", "")
        
        if request.platform == "prestashop":
            rows.append({
                "ID": product.get("id"),
                "Active (0/1)": "1",
                "Name*": name,
                "Categories (x,y,z...)": product.get("category", ""),
                "Price tax excl.": round(final_price, 2),
                "Tax rules ID": "1",
                "Wholesale price": product.get("price", 0),
                "On sale (0/1)": "0",
                "Discount amount": "0",
                "Discount percent": "0",
                "Discount from (yyyy-mm-dd)": "",
                "Discount to (yyyy-mm-dd)": "",
                "Reference #": product.get("sku", ""),
                "Supplier reference #": product.get("sku", ""),
                "Supplier": product.get("supplier_name", ""),
                "Manufacturer": product.get("brand", ""),
                "EAN13": product.get("ean", ""),
                "UPC": "",
                "Ecotax": "0",
                "Width": "",
                "Height": "",
                "Depth": "",
                "Weight": product.get("weight", ""),
                "Quantity": product.get("stock", 0),
                "Minimal quantity": "1",
                "Visibility": "both",
                "Additional shipping cost": "0",
                "Unity": "",
                "Unit price": "",
                "Short description": "",
                "Description": product.get("description", ""),
                "Tags (x,y,z...)": "",
                "Meta title": name,
                "Meta keywords": "",
                "Meta description": "",
                "URL rewritten": "",
                "Text when in stock": "En stock",
                "Text when backorder allowed": "",
                "Available for order (0 = No, 1 = Yes)": "1",
                "Product available date": "",
                "Product creation date": "",
                "Show price (0 = No, 1 = Yes)": "1",
                "Image URLs (x,y,z...)": product.get("image_url", ""),
                "Delete existing images (0 = No, 1 = Yes)": "0"
            })
        elif request.platform == "woocommerce":
            rows.append({
                "ID": "",
                "Type": "simple",
                "SKU": product.get("sku", ""),
                "Name": name,
                "Published": "1",
                "Is featured?": "0",
                "Visibility in catalog": "visible",
                "Short description": "",
                "Description": product.get("description", ""),
                "Date sale price starts": "",
                "Date sale price ends": "",
                "Tax status": "taxable",
                "Tax class": "",
                "In stock?": "1" if product.get("stock", 0) > 0 else "0",
                "Stock": product.get("stock", 0),
                "Backorders allowed?": "0",
                "Sold individually?": "0",
                "Weight (kg)": product.get("weight", ""),
                "Length (cm)": "",
                "Width (cm)": "",
                "Height (cm)": "",
                "Allow customer reviews?": "1",
                "Purchase note": "",
                "Sale price": "",
                "Regular price": round(final_price, 2),
                "Categories": product.get("category", ""),
                "Tags": "",
                "Shipping class": "",
                "Images": product.get("image_url", ""),
                "Download limit": "",
                "Download expiry days": "",
                "Parent": "",
                "Grouped products": "",
                "Upsells": "",
                "Cross-sells": "",
                "External URL": "",
                "Button text": "",
                "Position": "0",
                "Brands": product.get("brand", "")
            })
        elif request.platform == "shopify":
            rows.append({
                "Handle": product.get("sku", "").lower().replace(" ", "-"),
                "Title": name,
                "Body (HTML)": product.get("description", ""),
                "Vendor": product.get("brand", ""),
                "Product Category": product.get("category", ""),
                "Type": product.get("category", ""),
                "Tags": "",
                "Published": "TRUE",
                "Option1 Name": "Title",
                "Option1 Value": "Default Title",
                "Option2 Name": "",
                "Option2 Value": "",
                "Option3 Name": "",
                "Option3 Value": "",
                "Variant SKU": product.get("sku", ""),
                "Variant Grams": int(float(product.get("weight", 0) or 0) * 1000),
                "Variant Inventory Tracker": "shopify",
                "Variant Inventory Qty": product.get("stock", 0),
                "Variant Inventory Policy": "deny",
                "Variant Fulfillment Service": "manual",
                "Variant Price": round(final_price, 2),
                "Variant Compare At Price": "",
                "Variant Requires Shipping": "TRUE",
                "Variant Taxable": "TRUE",
                "Variant Barcode": product.get("ean", ""),
                "Image Src": product.get("image_url", ""),
                "Image Position": "1",
                "Image Alt Text": name,
                "Gift Card": "FALSE",
                "SEO Title": name,
                "SEO Description": product.get("description", "")[:160] if product.get("description") else "",
                "Google Shopping / Google Product Category": "",
                "Google Shopping / Gender": "",
                "Google Shopping / Age Group": "",
                "Google Shopping / MPN": product.get("sku", ""),
                "Google Shopping / AdWords Grouping": "",
                "Google Shopping / AdWords Labels": "",
                "Google Shopping / Condition": "new",
                "Google Shopping / Custom Product": "",
                "Google Shopping / Custom Label 0": "",
                "Google Shopping / Custom Label 1": "",
                "Google Shopping / Custom Label 2": "",
                "Google Shopping / Custom Label 3": "",
                "Google Shopping / Custom Label 4": "",
                "Variant Image": product.get("image_url", ""),
                "Variant Weight Unit": "kg",
                "Variant Tax Code": "",
                "Cost per item": product.get("price", 0),
                "Status": "active"
            })
    
    if not rows:
        raise HTTPException(status_code=400, detail="No hay productos para exportar")
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    
    output.seek(0)
    filename = f"catalog_{request.platform}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== DASHBOARD ENDPOINTS ====================

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    total_suppliers = await db.suppliers.count_documents({"user_id": user["id"]})
    total_products = await db.products.count_documents({"user_id": user["id"]})
    total_catalog_items = await db.catalog.count_documents({"user_id": user["id"]})
    low_stock_count = await db.products.count_documents({"user_id": user["id"], "stock": {"$gt": 0, "$lte": 5}})
    out_of_stock_count = await db.products.count_documents({"user_id": user["id"], "stock": 0})
    unread_notifications = await db.notifications.count_documents({"user_id": user["id"], "read": False})
    
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    recent_price_changes = await db.price_history.count_documents({"user_id": user["id"], "created_at": {"$gte": week_ago}})
    
    return DashboardStats(
        total_suppliers=total_suppliers,
        total_products=total_products,
        total_catalog_items=total_catalog_items,
        low_stock_count=low_stock_count,
        out_of_stock_count=out_of_stock_count,
        unread_notifications=unread_notifications,
        recent_price_changes=recent_price_changes
    )

@api_router.get("/dashboard/stock-alerts")
async def get_stock_alerts(user: dict = Depends(get_current_user)):
    low_stock = await db.products.find(
        {"user_id": user["id"], "stock": {"$gt": 0, "$lte": 5}},
        {"_id": 0, "user_id": 0}
    ).limit(10).to_list(10)
    
    out_of_stock = await db.products.find(
        {"user_id": user["id"], "stock": 0},
        {"_id": 0, "user_id": 0}
    ).limit(10).to_list(10)
    
    return {"low_stock": low_stock, "out_of_stock": out_of_stock}

# ==================== NOTIFICATIONS ENDPOINTS ====================

@api_router.get("/notifications", response_model=List[NotificationResponse])
async def get_notifications(
    unread_only: bool = False,
    skip: int = 0,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    query = {"user_id": user["id"]}
    if unread_only:
        query["read"] = False
    
    notifications = await db.notifications.find(query, {"_id": 0, "user_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return [NotificationResponse(**n) for n in notifications]

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(get_current_user)):
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": user["id"]},
        {"$set": {"read": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notificación no encontrada")
    return {"message": "Notificación marcada como leída"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(user: dict = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": user["id"]}, {"$set": {"read": True}})
    return {"message": "Todas las notificaciones marcadas como leídas"}

# ==================== PRICE HISTORY ENDPOINTS ====================

@api_router.get("/price-history", response_model=List[PriceHistoryResponse])
async def get_price_history(
    product_id: Optional[str] = None,
    days: int = 30,
    skip: int = 0,
    limit: int = 100,
    user: dict = Depends(get_current_user)
):
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    query = {"user_id": user["id"], "created_at": {"$gte": start_date}}
    if product_id:
        query["product_id"] = product_id
    
    history = await db.price_history.find(query, {"_id": 0, "user_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return [PriceHistoryResponse(**h) for h in history]

# ==================== WOOCOMMERCE INTEGRATION ENDPOINTS ====================

def get_woocommerce_client(config: dict) -> WooCommerceAPI:
    """Create WooCommerce API client from config"""
    return WooCommerceAPI(
        url=config['store_url'],
        consumer_key=config['consumer_key'],
        consumer_secret=config['consumer_secret'],
        version="wc/v3",
        timeout=30
    )

def mask_key(key: str) -> str:
    """Mask API key showing only last 4 characters"""
    if len(key) <= 4:
        return "****"
    return "*" * (len(key) - 4) + key[-4:]

@api_router.post("/woocommerce/configs", response_model=WooCommerceConfigResponse)
async def create_woocommerce_config(config: WooCommerceConfig, user: dict = Depends(get_current_user)):
    """Create a new WooCommerce store configuration"""
    config_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    config_doc = {
        "id": config_id,
        "user_id": user["id"],
        "name": config.name or "Mi Tienda WooCommerce",
        "store_url": config.store_url.rstrip('/'),
        "consumer_key": config.consumer_key,
        "consumer_secret": config.consumer_secret,
        "is_connected": False,
        "last_sync": None,
        "products_synced": 0,
        "created_at": now
    }
    
    await db.woocommerce_configs.insert_one(config_doc)
    
    return WooCommerceConfigResponse(
        id=config_id,
        name=config_doc["name"],
        store_url=config_doc["store_url"],
        consumer_key_masked=mask_key(config.consumer_key),
        is_connected=False,
        last_sync=None,
        products_synced=0,
        created_at=now
    )

@api_router.get("/woocommerce/configs", response_model=List[WooCommerceConfigResponse])
async def get_woocommerce_configs(user: dict = Depends(get_current_user)):
    """Get all WooCommerce configurations for the user"""
    configs = await db.woocommerce_configs.find(
        {"user_id": user["id"]}, 
        {"_id": 0, "consumer_secret": 0}
    ).to_list(100)
    
    return [
        WooCommerceConfigResponse(
            id=c["id"],
            name=c["name"],
            store_url=c["store_url"],
            consumer_key_masked=mask_key(c["consumer_key"]),
            is_connected=c.get("is_connected", False),
            last_sync=c.get("last_sync"),
            products_synced=c.get("products_synced", 0),
            created_at=c["created_at"]
        ) for c in configs
    ]

@api_router.get("/woocommerce/configs/{config_id}", response_model=WooCommerceConfigResponse)
async def get_woocommerce_config(config_id: str, user: dict = Depends(get_current_user)):
    """Get a specific WooCommerce configuration"""
    config = await db.woocommerce_configs.find_one(
        {"id": config_id, "user_id": user["id"]},
        {"_id": 0, "consumer_secret": 0}
    )
    if not config:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    
    return WooCommerceConfigResponse(
        id=config["id"],
        name=config["name"],
        store_url=config["store_url"],
        consumer_key_masked=mask_key(config["consumer_key"]),
        is_connected=config.get("is_connected", False),
        last_sync=config.get("last_sync"),
        products_synced=config.get("products_synced", 0),
        created_at=config["created_at"]
    )

@api_router.put("/woocommerce/configs/{config_id}", response_model=WooCommerceConfigResponse)
async def update_woocommerce_config(config_id: str, update: WooCommerceConfigUpdate, user: dict = Depends(get_current_user)):
    """Update a WooCommerce configuration"""
    existing = await db.woocommerce_configs.find_one({"id": config_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if "store_url" in update_data:
        update_data["store_url"] = update_data["store_url"].rstrip('/')
    
    if update_data:
        await db.woocommerce_configs.update_one({"id": config_id}, {"$set": update_data})
    
    updated = await db.woocommerce_configs.find_one({"id": config_id}, {"_id": 0, "consumer_secret": 0})
    return WooCommerceConfigResponse(
        id=updated["id"],
        name=updated["name"],
        store_url=updated["store_url"],
        consumer_key_masked=mask_key(updated["consumer_key"]),
        is_connected=updated.get("is_connected", False),
        last_sync=updated.get("last_sync"),
        products_synced=updated.get("products_synced", 0),
        created_at=updated["created_at"]
    )

@api_router.delete("/woocommerce/configs/{config_id}")
async def delete_woocommerce_config(config_id: str, user: dict = Depends(get_current_user)):
    """Delete a WooCommerce configuration"""
    result = await db.woocommerce_configs.delete_one({"id": config_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    return {"message": "Configuración eliminada"}

@api_router.post("/woocommerce/configs/{config_id}/test")
async def test_woocommerce_connection(config_id: str, user: dict = Depends(get_current_user)):
    """Test connection to WooCommerce store"""
    config = await db.woocommerce_configs.find_one({"id": config_id, "user_id": user["id"]})
    if not config:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    
    try:
        wcapi = get_woocommerce_client(config)
        # Test connection by getting store info
        response = await asyncio.to_thread(wcapi.get, "")
        
        if response.status_code == 200:
            await db.woocommerce_configs.update_one(
                {"id": config_id},
                {"$set": {"is_connected": True}}
            )
            store_info = response.json()
            return {
                "status": "success",
                "message": "Conexión exitosa",
                "store_name": store_info.get("name", "Tienda WooCommerce"),
                "store_description": store_info.get("description", "")
            }
        else:
            await db.woocommerce_configs.update_one(
                {"id": config_id},
                {"$set": {"is_connected": False}}
            )
            return {
                "status": "error",
                "message": f"Error de conexión: {response.status_code}"
            }
    except Exception as e:
        await db.woocommerce_configs.update_one(
            {"id": config_id},
            {"$set": {"is_connected": False}}
        )
        return {
            "status": "error",
            "message": f"Error de conexión: {str(e)}"
        }

@api_router.post("/woocommerce/export", response_model=WooCommerceExportResult)
async def export_to_woocommerce(request: WooCommerceExportRequest, user: dict = Depends(get_current_user)):
    """Export catalog products to WooCommerce"""
    # Get WooCommerce config
    config = await db.woocommerce_configs.find_one({"id": request.config_id, "user_id": user["id"]})
    if not config:
        raise HTTPException(status_code=404, detail="Configuración de WooCommerce no encontrada")
    
    # Get catalog items to export
    query = {"user_id": user["id"], "active": True}
    if request.catalog_ids:
        query["id"] = {"$in": request.catalog_ids}
    
    catalog_items = await db.catalog.find(query, {"_id": 0}).to_list(1000)
    
    if not catalog_items:
        return WooCommerceExportResult(status="warning", errors=["No hay productos activos para exportar"])
    
    # Get product details for each catalog item
    product_ids = [item["product_id"] for item in catalog_items]
    products = await db.products.find({"id": {"$in": product_ids}}, {"_id": 0}).to_list(1000)
    products_map = {p["id"]: p for p in products}
    
    # Create WooCommerce client
    wcapi = get_woocommerce_client(config)
    
    created = 0
    updated = 0
    failed = 0
    errors = []
    
    # Get existing products by SKU for update check
    existing_skus = {}
    if request.update_existing:
        try:
            response = await asyncio.to_thread(wcapi.get, "products", params={"per_page": 100})
            if response.status_code == 200:
                for p in response.json():
                    if p.get("sku"):
                        existing_skus[p["sku"]] = p["id"]
        except Exception as e:
            logger.warning(f"Could not fetch existing products: {e}")
    
    # Export each product
    for catalog_item in catalog_items:
        product = products_map.get(catalog_item["product_id"])
        if not product:
            failed += 1
            errors.append(f"Producto no encontrado: {catalog_item['product_id']}")
            continue
        
        try:
            # Prepare WooCommerce product data
            wc_product = {
                "name": catalog_item.get("custom_name") or product.get("name", "Producto sin nombre"),
                "type": "simple",
                "regular_price": str(catalog_item.get("final_price", product.get("price", 0))),
                "description": product.get("description", ""),
                "short_description": product.get("short_description", ""),
                "sku": product.get("sku", ""),
                "manage_stock": True,
                "stock_quantity": product.get("stock", 0),
                "categories": [],
                "images": []
            }
            
            # Add category if exists
            if product.get("category"):
                wc_product["categories"] = [{"name": product["category"]}]
            
            # Add images if exist
            for img_field in ["image_url", "image_url2", "image_url3"]:
                if product.get(img_field):
                    wc_product["images"].append({"src": product[img_field]})
            
            # Add weight if exists
            if product.get("weight"):
                wc_product["weight"] = str(product["weight"])
            
            # Check if product exists (by SKU) and update or create
            sku = product.get("sku", "")
            if sku and sku in existing_skus and request.update_existing:
                # Update existing product
                response = await asyncio.to_thread(
                    wcapi.put, 
                    f"products/{existing_skus[sku]}", 
                    wc_product
                )
                if response.status_code in [200, 201]:
                    updated += 1
                else:
                    failed += 1
                    errors.append(f"Error actualizando {sku}: {response.text[:100]}")
            else:
                # Create new product
                response = await asyncio.to_thread(wcapi.post, "products", wc_product)
                if response.status_code in [200, 201]:
                    created += 1
                else:
                    failed += 1
                    errors.append(f"Error creando {sku or product.get('name', 'producto')}: {response.text[:100]}")
                    
        except Exception as e:
            failed += 1
            errors.append(f"Error procesando {product.get('sku', 'producto')}: {str(e)[:100]}")
    
    # Update sync stats
    now = datetime.now(timezone.utc).isoformat()
    await db.woocommerce_configs.update_one(
        {"id": request.config_id},
        {"$set": {
            "last_sync": now,
            "products_synced": created + updated
        }}
    )
    
    # Create notification
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "type": "woocommerce_export",
        "message": f"Exportación WooCommerce: {created} creados, {updated} actualizados, {failed} errores",
        "product_id": None,
        "product_name": None,
        "user_id": user["id"],
        "read": False,
        "created_at": now
    })
    
    return WooCommerceExportResult(
        status="success" if failed == 0 else "partial" if (created + updated) > 0 else "error",
        created=created,
        updated=updated,
        failed=failed,
        errors=errors[:10]  # Limit to 10 errors
    )

@api_router.get("/woocommerce/configs/{config_id}/products")
async def get_woocommerce_products(config_id: str, page: int = 1, per_page: int = 20, user: dict = Depends(get_current_user)):
    """Get products from WooCommerce store"""
    config = await db.woocommerce_configs.find_one({"id": config_id, "user_id": user["id"]})
    if not config:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    
    try:
        wcapi = get_woocommerce_client(config)
        response = await asyncio.to_thread(
            wcapi.get, 
            "products", 
            params={"page": page, "per_page": per_page}
        )
        
        if response.status_code == 200:
            products = response.json()
            total = response.headers.get('X-WP-Total', 0)
            return {
                "products": products,
                "total": int(total),
                "page": page,
                "per_page": per_page
            }
        else:
            return {"status": "error", "message": f"Error: {response.status_code}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ==================== HEALTH CHECK ====================

@api_router.get("/health")
async def health_check():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include router and configure app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
