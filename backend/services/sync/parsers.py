import csv
import io
import logging
import os
import zipfile

import xlrd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _sanitize_csv_cell(value):
    """Previene CSV formula injection eliminando prefijos peligrosos."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


def parse_csv_content(content: bytes) -> list:
    try:
        decoded = content.decode('utf-8')
    except Exception:
        decoded = content.decode('latin-1')
    reader = csv.DictReader(io.StringIO(decoded))
    result = []
    for row in reader:
        sanitized = {k: _sanitize_csv_cell(v) if isinstance(v, str) else v
                     for k, v in row.items()}
        result.append(sanitized)
    return result


def parse_xlsx_content(content: bytes) -> list:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).lower().strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(row)]


def parse_xls_content(content: bytes) -> list:
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, c)).lower().strip() for c in range(ws.ncols)]
    return [{headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)} for r in range(1, ws.nrows)]


def parse_xml_content(content: bytes) -> list:
    """Parse XML content safely - prevents XXE attacks

    SECURITY FIX: Disable external entities to prevent XXE injection
    """
    try:
        decoded = content.decode('utf-8')
    except Exception:
        decoded = content.decode('latin-1')

    # SECURITY: Use defusedxml to prevent XXE attacks
    # Disables: external entities, internal DTD parsing, entity expansion
    try:
        from defusedxml import xmltodict as safe_xmltodict
        data = safe_xmltodict.parse(decoded, disable_entities=True, process_namespaces=False)
    except ImportError:
        raise ImportError(
            "El parseo de XML requiere la librería defusedxml para protección XXE. "
            "Instalar con: pip install defusedxml"
        )

    for key in ['products', 'items', 'catalog', 'data', 'root']:
        if key in data:
            items = data[key]
            if isinstance(items, dict):
                for subkey in items:
                    if isinstance(items[subkey], list):
                        return items[subkey]
            elif isinstance(items, list):
                return items
    return []


def _detect_best_separator(first_line: str, preferred: str) -> str:
    """Return the separator that produces the most columns in first_line.
    Falls back to preferred if no alternative wins by a clear margin."""
    candidates = [preferred, ';', ',', '\t', '|']
    # deduplicate while preserving order
    seen = set()
    candidates = [c for c in candidates if not (c in seen or seen.add(c))]
    best_sep = preferred
    best_count = len(list(csv.reader([first_line], delimiter=preferred, quotechar='"'))[0]) if first_line else 1
    for sep in candidates[1:]:
        try:
            count = len(list(csv.reader([first_line], delimiter=sep, quotechar='"'))[0])
        except Exception:
            continue
        if count > best_count:
            best_count = count
            best_sep = sep
    if best_sep != preferred:
        logger.info(f"Auto-detected separator {repr(best_sep)} ({best_count} cols) instead of {repr(preferred)}")
    return best_sep


def parse_text_file(content: bytes, separator: str = ";", header_row: int = 1) -> list:
    """Parse a text file (CSV/TXT). header_row=0 means no header (positional col names).
    If the configured separator produces only 1 column, auto-detects the best separator."""
    try:
        decoded = content.decode('utf-8-sig', errors='replace')
    except Exception:
        try:
            decoded = content.decode('utf-8', errors='replace')
        except Exception:
            decoded = content.decode('latin-1', errors='replace')
    lines = decoded.strip().split('\n')
    if header_row > 1:
        lines = lines[header_row - 1:]
    if not lines:
        return []
    if separator == '\\t':
        separator = '\t'
    first_line = lines[0].rstrip('\r') if lines else ''
    # Auto-detect separator when configured one yields only 1 column
    separator = _detect_best_separator(first_line, separator)
    if header_row == 0:
        first_row_parsed = list(csv.reader([first_line], delimiter=separator, quotechar='"'))
        num_cols = len(first_row_parsed[0]) if first_row_parsed else 0
        fieldnames = [f'col_{i}' for i in range(num_cols)]
        reader = csv.DictReader(lines, fieldnames=fieldnames, delimiter=separator, quotechar='"')
    else:
        reader = csv.DictReader(lines, delimiter=separator, quotechar='"')
    return list(reader)


def extract_zip_files(content: bytes) -> dict:
    """Extract all files from a ZIP archive, returns {filename: bytes}.
    Valida paths para prevenir path traversal (Zip Slip)."""
    result = {}
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            normalized = os.path.normpath(info.filename)
            if normalized.startswith('..') or os.path.isabs(normalized):
                logger.warning(f"Zip Slip: path sospechoso ignorado: {info.filename}")
                continue
            with zf.open(info.filename) as f:
                result[info.filename] = f.read()
    return result
