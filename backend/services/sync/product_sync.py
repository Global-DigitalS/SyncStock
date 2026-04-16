"""
Orquestacion de sincronizacion de proveedores: descarga, parseo, normalizacion y upsert en BD.
"""
import asyncio
import csv
import io
import logging
import os
import uuid
from datetime import UTC, datetime

import xlrd
from openpyxl import load_workbook
from pymongo import UpdateOne

from config import LOW_STOCK_THRESHOLD, PRICE_CHANGE_THRESHOLD_PERCENT
from services.database import db
from services.sku_cache import SKUCache
from services.sync.downloaders import download_file_from_ftp, download_file_from_url
from services.sync.normalizer import apply_column_mapping, normalize_product_data
from services.sync.notifications import send_realtime_notification, send_sync_progress
from services.sync.parsers import (
    _detect_best_separator,
    extract_zip_files,
    parse_text_file,
    parse_xml_content,
)

logger = logging.getLogger(__name__)

BULK_BATCH_SIZE = 500
PROGRESS_REPORT_INTERVAL = 2000


async def prefetch_existing_products(supplier_id: str, user_id: str) -> SKUCache:
    """Crea e inicializa un SKUCache para busquedas eficientes durante la sincronizacion."""
    return SKUCache(supplier_id, user_id)


async def bulk_upsert_products(supplier: dict, normalized_products: list, sku_cache: SKUCache, now: str) -> dict:
    """
    Bulk upsert de productos usando operaciones en batch y cache SKU para lookups eficientes.

    Optimizado para 1M+ productos:
    - Usa UpdateOne con upsert=True para evitar errores E11000 de clave duplicada
    - Procesa en chunks para evitar cargar 1M productos en memoria
    - Tamano de batch 5000 para mejor rendimiento
    """
    imported = 0
    updated = 0
    errors = 0

    supplier_id = supplier['id']
    supplier_name = supplier['name']
    user_id = supplier['user_id']
    total = len(normalized_products)

    CHUNK_SIZE = 5000
    DB_BATCH_SIZE = 5000

    product_ops = []
    price_history_docs = []
    notification_docs = []

    first_batch_skus = [
        p.get('sku') for p in normalized_products[:CHUNK_SIZE]
        if p.get('sku')
    ]
    if first_batch_skus:
        await sku_cache.populate_batch(first_batch_skus)

    for i, normalized in enumerate(normalized_products):
        try:
            sku = normalized.get('sku')
            name = normalized.get('name', '')
            if not sku or not name:
                errors += 1
                continue

            product_doc = {
                "sku": sku, "name": name,
                "description": normalized.get('description'),
                "price": normalized.get('price', 0),
                "stock": normalized.get('stock', 0),
                "category": normalized.get('category'),
                "brand": normalized.get('brand'),
                "ean": normalized.get('ean'),
                "weight": normalized.get('weight'),
                "image_url": normalized.get('image_url'),
                "supplier_id": supplier_id,
                "supplier_name": supplier_name,
                "user_id": user_id,
                "updated_at": now
            }

            existing = sku_cache.get(sku)

            if existing:
                old_price = existing.price
                new_price = product_doc['price']
                if old_price != new_price and old_price > 0:
                    change_pct = ((new_price - old_price) / old_price) * 100
                    price_history_docs.append({
                        "id": str(uuid.uuid4()), "product_id": existing.id,
                        "product_name": name, "old_price": old_price,
                        "new_price": new_price, "change_percentage": change_pct,
                        "user_id": user_id, "created_at": now
                    })
                    if abs(change_pct) >= PRICE_CHANGE_THRESHOLD_PERCENT:
                        direction = "subido" if change_pct > 0 else "bajado"
                        notification_docs.append({
                            "id": str(uuid.uuid4()), "type": "price_change",
                            "message": f"Precio de '{name[:40]}' ha {direction} {abs(change_pct):.1f}% ({old_price:.2f}EUR -> {new_price:.2f}EUR)",
                            "product_id": existing.id, "product_name": name,
                            "user_id": user_id, "read": False, "created_at": now
                        })

                old_stock = existing.stock
                new_stock = product_doc['stock']
                if old_stock > 0 and new_stock == 0:
                    notification_docs.append({
                        "id": str(uuid.uuid4()), "type": "stock_out",
                        "message": f"Producto '{name[:40]}' sin stock",
                        "product_id": existing.id, "product_name": name,
                        "user_id": user_id, "read": False, "created_at": now
                    })
                elif old_stock > LOW_STOCK_THRESHOLD and 0 < new_stock <= LOW_STOCK_THRESHOLD:
                    notification_docs.append({
                        "id": str(uuid.uuid4()), "type": "stock_low",
                        "message": f"Producto '{name[:40]}' con stock bajo ({new_stock} uds)",
                        "product_id": existing.id, "product_name": name,
                        "user_id": user_id, "read": False, "created_at": now
                    })

                product_ops.append(UpdateOne(
                    {"supplier_id": supplier_id, "sku": sku, "id": existing.id},
                    {"$set": product_doc, "$setOnInsert": {"id": existing.id, "created_at": now}},
                    upsert=False
                ))
                product_ops.append(UpdateOne(
                    {"supplier_id": supplier_id, "sku": sku, "id": {"$ne": existing.id}},
                    {"$set": product_doc, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now}},
                    upsert=True
                ))
                updated += 1
            else:
                product_ops.append(UpdateOne(
                    {"supplier_id": supplier_id, "sku": sku},
                    {"$set": product_doc, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now}},
                    upsert=True
                ))
                imported += 1

            if len(product_ops) >= DB_BATCH_SIZE:
                await db.products.bulk_write(product_ops, ordered=False)
                product_ops = []

            if len(price_history_docs) >= DB_BATCH_SIZE:
                await db.price_history.insert_many(price_history_docs, ordered=False)
                price_history_docs = []

            if len(notification_docs) >= DB_BATCH_SIZE:
                await db.notifications.insert_many(notification_docs, ordered=False)
                notification_docs = []

            if (i + 1) % CHUNK_SIZE == 0:
                next_batch_skus = [
                    p.get('sku') for p in normalized_products[i+1:i+1+CHUNK_SIZE]
                    if p.get('sku')
                ]
                if next_batch_skus:
                    await sku_cache.populate_batch(next_batch_skus)
                await send_sync_progress(user_id, supplier_name, i + 1, total)

        except Exception as e:
            logger.error(f"Error processing product: {e}")
            errors += 1

    if product_ops:
        await db.products.bulk_write(product_ops, ordered=False)
    if price_history_docs:
        await db.price_history.insert_many(price_history_docs, ordered=False)
    if notification_docs:
        await db.notifications.insert_many(notification_docs, ordered=False)

    cache_stats = sku_cache.get_stats()
    logger.info(f"Sync complete for {supplier_name}: Cache stats: {cache_stats}")

    return {"imported": imported, "updated": updated, "errors": errors}


async def process_supplier_file(supplier: dict, content: bytes) -> dict:
    file_format = supplier.get('file_format', 'csv').lower()
    separator = supplier.get('csv_separator', ';')
    enclosure = supplier.get('csv_enclosure', '"')
    _header_row_raw = supplier.get('csv_header_row', 1)
    header_row = 1 if _header_row_raw is None else int(_header_row_raw)
    column_mapping = supplier.get('column_mapping')
    strip_ean_quotes = supplier.get('strip_ean_quotes', False)
    detected_columns = []

    # Auto-detectar ZIP por magic bytes (firma PK)
    if len(content) >= 4 and content[:2] == b'PK':
        file_format = 'zip'

    try:
        if file_format == 'zip':
            try:
                extracted = extract_zip_files(content)
            except Exception as e:
                return {"imported": 0, "updated": 0, "errors": 0, "message": f"Error al descomprimir ZIP: {e}"}

            if not extracted:
                return {"imported": 0, "updated": 0, "errors": 0, "message": "El archivo ZIP esta vacio"}

            logger.info(f"ZIP detectado con {len(extracted)} archivo(s): {list(extracted.keys())}")

            compatible_exts = ('.csv', '.txt', '.xlsx', '.xls', '.xml')
            compatible_files = [
                (fname, fcontent)
                for fname, fcontent in extracted.items()
                if fname.lower().endswith(compatible_exts)
                and not fname.split('/')[-1].startswith('.')
                and not fname.split('/')[-1].startswith('__')
            ]

            if not compatible_files:
                names = list(extracted.keys())
                return {
                    "imported": 0, "updated": 0, "errors": 0,
                    "message": f"El ZIP no contiene archivos compatibles (csv/xlsx/xls/xml/txt). Archivos encontrados: {names}"
                }

            if len(compatible_files) == 1:
                best_fname, best_content = compatible_files[0]
                best_fmt = 'csv' if best_fname.lower().endswith(('.csv', '.txt')) else best_fname.rsplit('.', 1)[-1].lower()
                logger.info(f"ZIP (unico archivo): procesando '{best_fname}' como {best_fmt.upper()}")
                zip_supplier = {**supplier, 'file_format': best_fmt}
                return await process_supplier_file(zip_supplier, best_content)

            logger.info(f"ZIP multi-archivo: detectando roles automaticamente para {len(compatible_files)} archivos")
            role_keywords = {
                'stock':    ['stock', 'inventory', 'disponibilidad', 'existencias', 'qty', 'quantity'],
                'prices':   ['price', 'precio', 'tarif', 'coste', 'cost', 'pvp', 'pvd'],
                'products': ['product', 'catalog', 'article', 'catalogo', 'articulo', 'master', 'items'],
            }

            all_file_data = {}
            for fname, fcontent in compatible_files:
                fname_base = fname.split('/')[-1].lower()
                detected_role = 'products'
                for role, keywords in role_keywords.items():
                    if any(kw in fname_base for kw in keywords):
                        detected_role = role
                        break

                file_hdr = header_row
                try:
                    if fname.lower().endswith(('.xlsx', '.xls')):
                        fmt = 'xlsx' if fname.lower().endswith('.xlsx') else 'xls'
                        if fmt == 'xlsx':
                            wb = load_workbook(filename=io.BytesIO(fcontent), read_only=True)
                            ws = wb.active
                            rows = list(ws.iter_rows(values_only=True))
                            hdrs = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
                            file_rows = [dict(zip(hdrs, r)) for r in rows[1:] if any(r)]
                        else:
                            wb = xlrd.open_workbook(file_contents=fcontent)
                            ws = wb.sheet_by_index(0)
                            hdrs = [str(ws.cell_value(0, c)).strip() or f'col_{c}' for c in range(ws.ncols)]
                            file_rows = [{hdrs[c]: ws.cell_value(r, c) for c in range(ws.ncols)} for r in range(1, ws.nrows)]
                    else:
                        file_rows = parse_text_file(fcontent, separator, file_hdr)
                except Exception as fe:
                    logger.warning(f"  No se pudo parsear {fname}: {fe}")
                    continue

                if detected_role not in all_file_data or len(file_rows) > len(all_file_data[detected_role]):
                    all_file_data[detected_role] = file_rows
                logger.info(f"  {fname} -> role={detected_role}, filas={len(file_rows)}")

            if not all_file_data:
                return {"imported": 0, "updated": 0, "errors": 0, "message": "No se pudo parsear ningun archivo del ZIP"}

            products_data = all_file_data.get('products', [])
            prices_data   = all_file_data.get('prices', [])
            stock_data    = all_file_data.get('stock', [])

            if not products_data:
                largest_role = max(all_file_data, key=lambda r: len(all_file_data[r]))
                products_data = all_file_data[largest_role]

            prices_merge_key = None
            prices_lookup = {}
            if prices_data:
                sample_product = products_data[0] if products_data else {}
                product_keys = list(sample_product.keys())
                common_merge_key = None
                for pk in product_keys:
                    if pk in prices_data[0]:
                        common_merge_key = pk
                        break
                prices_merge_key = common_merge_key or list(prices_data[0].keys())[0]
                logger.info(f"ZIP prices merge_key: {prices_merge_key} (common={common_merge_key is not None})")
                if prices_merge_key:
                    for row in prices_data:
                        k = str(row.get(prices_merge_key, '')).strip()
                        if k:
                            prices_lookup[k] = row

            stock_merge_key = None
            stock_lookup = {}
            if stock_data:
                sample_product = products_data[0] if products_data else {}
                product_keys = list(sample_product.keys())
                common_merge_key = None
                for pk in product_keys:
                    if pk in stock_data[0]:
                        common_merge_key = pk
                        break
                stock_merge_key = common_merge_key or list(stock_data[0].keys())[0]
                logger.info(f"ZIP stock merge_key: {stock_merge_key} (common={common_merge_key is not None})")
                if stock_merge_key:
                    for row in stock_data:
                        k = str(row.get(stock_merge_key, '')).strip()
                        if k:
                            stock_lookup[k] = row

            logger.info(f"ZIP merge: {len(products_data)} productos, {len(prices_lookup)} precios (key={prices_merge_key}), {len(stock_lookup)} stock (key={stock_merge_key})")

            zip_detected_cols = []
            if products_data:
                sample = dict(products_data[0])
                sample_id = str(sample.get(list(sample.keys())[0], '')).strip()
                if prices_merge_key and sample_id in prices_lookup:
                    for k in prices_lookup[sample_id]:
                        if header_row == 0:
                            if k != prices_merge_key:
                                sample[f"prices_{k}"] = prices_lookup[sample_id][k]
                        elif k not in sample:
                            sample[k] = prices_lookup[sample_id][k]
                if stock_merge_key and sample_id in stock_lookup:
                    for k in stock_lookup[sample_id]:
                        if header_row == 0:
                            if k != stock_merge_key:
                                sample[f"stock_{k}"] = stock_lookup[sample_id][k]
                        elif k not in sample:
                            sample[k] = stock_lookup[sample_id][k]
                zip_detected_cols = list(sample.keys())
                await db.suppliers.update_one(
                    {"id": supplier['id']},
                    {"$set": {"detected_columns": zip_detected_cols}}
                )

            now = datetime.now(UTC).isoformat()
            needs_mapping = False
            errors = 0
            merge_key = list(products_data[0].keys())[0] if products_data else None
            column_mapping = supplier.get('column_mapping')

            normalized_products = []
            for raw in products_data:
                try:
                    prod_id = str(raw.get(merge_key, '')).strip() if merge_key else ''
                    merged = dict(raw)
                    if prod_id and prod_id in prices_lookup:
                        for k, v in prices_lookup[prod_id].items():
                            if header_row == 0:
                                if k != prices_merge_key:
                                    merged[f"prices_{k}"] = v
                            elif k not in merged or not merged[k]:
                                merged[k] = v
                    if prod_id and prod_id in stock_lookup:
                        for k, v in stock_lookup[prod_id].items():
                            if header_row == 0:
                                if k != stock_merge_key:
                                    merged[f"stock_{k}"] = v
                            elif k not in merged or not merged[k]:
                                merged[k] = v

                    normalized = apply_column_mapping(merged, column_mapping, strip_ean_quotes) if column_mapping else normalize_product_data(merged, strip_ean_quotes)
                    sku = normalized.get('sku') or prod_id
                    name = normalized.get('name', '')
                    if not sku or not name:
                        errors += 1
                        needs_mapping = True
                        continue
                    normalized_products.append(normalized)
                except Exception as e:
                    logger.error(f"Error procesando producto ZIP: {e}")
                    errors += 1

            sku_cache = await prefetch_existing_products(supplier['id'], supplier['user_id'])
            bulk_result = await bulk_upsert_products(supplier, normalized_products, sku_cache, now)
            imported = bulk_result["imported"]
            updated = bulk_result["updated"]
            errors += bulk_result["errors"]

            msg = f"ZIP procesado: {imported} importados, {updated} actualizados"
            result = {"imported": imported, "updated": updated, "errors": errors, "message": msg,
                      "needs_mapping": needs_mapping, "detected_columns": zip_detected_cols}
            if needs_mapping and imported + updated == 0:
                result["status"] = "needs_mapping"
                result["message"] = (
                    f"El ZIP se proceso pero no se importaron productos. "
                    f"Columnas disponibles: {', '.join(zip_detected_cols[:8])}... "
                    f"Configura el mapeo de columnas o re-aplica la plantilla del proveedor."
                )
            return result

        elif file_format == 'csv':
            try:
                decoded = content.decode('utf-8-sig')
            except Exception:
                try:
                    decoded = content.decode('utf-8')
                except Exception:
                    decoded = content.decode('latin-1')
            lines = decoded.split('\n')
            if header_row > 1:
                lines = lines[header_row-1:]
            if separator == '\\t':
                separator = '\t'
            first_line_raw = lines[0].rstrip('\r') if lines else ''
            separator = _detect_best_separator(first_line_raw, separator)
            if header_row == 0:
                first_line = first_line_raw
                first_row_parsed = list(csv.reader([first_line], delimiter=separator, quotechar=enclosure if enclosure else '"'))
                num_cols = len(first_row_parsed[0]) if first_row_parsed else 0
                fieldnames = [f'col_{i}' for i in range(num_cols)]
                reader = csv.DictReader(lines, fieldnames=fieldnames, delimiter=separator, quotechar=enclosure if enclosure else '"')
            else:
                reader = csv.DictReader(lines, delimiter=separator, quotechar=enclosure if enclosure else '"')
            raw_products = list(reader)
            if raw_products:
                detected_columns = list(raw_products[0].keys())
        elif file_format in ['xlsx', 'xls']:
            if file_format == 'xlsx':
                wb = load_workbook(filename=io.BytesIO(content), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            else:
                wb = xlrd.open_workbook(file_contents=content)
                ws = wb.sheet_by_index(0)
                rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
            if not rows:
                return {"imported": 0, "updated": 0, "errors": 0}
            if header_row > 1:
                rows = rows[header_row-1:]
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
            detected_columns = headers
            raw_products = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
        elif file_format == 'xml':
            raw_products = parse_xml_content(content)
            if raw_products:
                detected_columns = list(raw_products[0].keys())
        else:
            return {"imported": 0, "updated": 0, "errors": 0, "message": "Unsupported format"}

        if detected_columns:
            await db.suppliers.update_one({"id": supplier['id']}, {"$set": {"detected_columns": detected_columns}})

        now = datetime.now(UTC).isoformat()
        imported = 0
        updated = 0
        errors = 0
        needs_mapping = False

        if not column_mapping and detected_columns:
            col_names_lower = [c.lower().strip().replace(' ', '_').replace('-', '_') for c in detected_columns]
            col_names_simple = [c.lower().strip() for c in detected_columns]
            all_col_variants = col_names_lower + col_names_simple

            sku_aliases = ['sku', 'codigo', 'referencia', 'ref', 'reference', 'id', 'cod', 'articulo',
                          'codigo_articulo', 'product_id', 'item_code', 'code', 'producto']
            name_aliases = ['name', 'nombre', 'title', 'titulo', 'product', 'producto', 'descripcion',
                           'description', 'denominacion', 'item', 'articulo', 'desc']

            has_sku = any(alias in all_col_variants for alias in sku_aliases)
            has_name = any(alias in all_col_variants for alias in name_aliases)

            if not has_sku or not has_name:
                needs_mapping = True
                logger.warning(f"Supplier {supplier.get('name')}: needs_mapping=True. Columns detected: {detected_columns[:10]}. has_sku={has_sku}, has_name={has_name}")

        normalized_products = []
        for raw in raw_products:
            try:
                normalized = apply_column_mapping(raw, column_mapping, strip_ean_quotes) if column_mapping else normalize_product_data(raw, strip_ean_quotes)
                if not normalized.get('sku') or not normalized.get('name'):
                    errors += 1
                    continue
                normalized_products.append(normalized)
            except Exception as e:
                logger.error(f"Error processing product: {e}")
                errors += 1

        sku_cache = await prefetch_existing_products(supplier['id'], supplier['user_id'])
        bulk_result = await bulk_upsert_products(supplier, normalized_products, sku_cache, now)
        imported = bulk_result["imported"]
        updated = bulk_result["updated"]
        errors += bulk_result["errors"]

        result = {"imported": imported, "updated": updated, "errors": errors, "detected_columns": detected_columns}

        if needs_mapping:
            result["needs_mapping"] = True
            if errors > 0 and (imported + updated) == 0:
                result["message"] = f"Las columnas detectadas ({', '.join(detected_columns[:5])}...) no coinciden con los nombres estandar. Configura el mapeo de columnas para este proveedor."
                result["status"] = "needs_mapping"
            else:
                result["message"] = "Importacion parcial. Se recomienda configurar el mapeo de columnas para mejores resultados."
        elif errors > 0 and (imported + updated) == 0:
            result["message"] = f"No se pudieron importar productos. Verifica el formato del archivo y las columnas: {', '.join(detected_columns[:5])}"
            result["status"] = "error"

        return result
    except Exception as e:
        logger.error(f"Error processing file: {e}")
        return {"imported": 0, "updated": 0, "errors": 1, "message": str(e), "detected_columns": []}


async def record_sync_history(supplier: dict, result: dict, sync_type: str, duration: float, error_message: str = None):
    """Registrar historial de sincronizacion"""
    status = "success" if result.get("status") == "success" else "error" if result.get("status") == "error" else "partial"
    if result.get("errors", 0) > 0 and result.get("imported", 0) + result.get("updated", 0) > 0:
        status = "partial"

    await db.sync_history.insert_one({
        "id": str(uuid.uuid4()),
        "supplier_id": supplier["id"],
        "supplier_name": supplier["name"],
        "sync_type": sync_type,
        "status": status,
        "imported": result.get("imported", 0),
        "updated": result.get("updated", 0),
        "errors": result.get("errors", 0),
        "duration_seconds": round(duration, 2),
        "error_message": error_message or result.get("message"),
        "user_id": supplier["user_id"],
        "created_at": datetime.now(UTC).isoformat()
    })


async def sync_supplier(supplier: dict, sync_type: str = "manual") -> dict:
    connection_type = supplier.get('connection_type', 'ftp')
    if connection_type == 'url':
        if not supplier.get('file_url'):
            return {"status": "skipped", "message": "URL no configurada"}
    else:
        if not supplier.get('ftp_host') or not supplier.get('ftp_path'):
            return {"status": "skipped", "message": "FTP no configurado"}

    start_time = datetime.now(UTC)
    try:
        logger.info(f"Syncing supplier: {supplier['name']} (via {connection_type})")
        if connection_type == 'url':
            from services.encryption import decrypt_password
            url_username = supplier.get('url_username')
            url_password = supplier.get('url_password')
            if url_password:
                url_password = decrypt_password(url_password)
            content = await download_file_from_url(supplier['file_url'], url_username, url_password)
        else:
            content = await download_file_from_ftp(supplier)
        result = await process_supplier_file(supplier, content)
        now = datetime.now(UTC)
        duration = (now - start_time).total_seconds()

        product_count = await db.products.count_documents({"supplier_id": supplier['id']})
        await db.suppliers.update_one(
            {"id": supplier['id']},
            {"$set": {"product_count": product_count, "last_sync": now.isoformat()}}
        )

        notification = {
            "id": str(uuid.uuid4()), "type": "sync_complete",
            "message": f"Sincronizacion completada: {supplier['name']} - {result['imported']} nuevos, {result['updated']} actualizados",
            "product_id": None, "product_name": None,
            "user_id": supplier["user_id"], "read": False, "created_at": now.isoformat()
        }
        await db.notifications.insert_one(notification)
        await send_realtime_notification(supplier["user_id"], notification)

        final_result = {"status": "success", **result}
        await record_sync_history(supplier, final_result, sync_type, duration)
        logger.info(f"Sync complete for {supplier['name']}: {result}")
        return final_result
    except Exception as e:
        duration = (datetime.now(UTC) - start_time).total_seconds()
        logger.error(f"Error syncing supplier {supplier['name']}: {e}")

        notification = {
            "id": str(uuid.uuid4()), "type": "sync_error",
            "message": f"Error en sincronizacion: {supplier['name']} - {str(e)[:100]}",
            "product_id": None, "product_name": None,
            "user_id": supplier["user_id"], "read": False,
            "created_at": datetime.now(UTC).isoformat()
        }
        await db.notifications.insert_one(notification)
        await send_realtime_notification(supplier["user_id"], notification)

        error_result = {"status": "error", "message": str(e), "imported": 0, "updated": 0, "errors": 1}
        await record_sync_history(supplier, error_result, sync_type, duration, str(e))
        return error_result


async def sync_all_suppliers():
    from services.unified_sync import _global_sync_lock
    logger.info("Starting scheduled sync for all suppliers (waiting for global lock)...")
    async with _global_sync_lock:
        ftp_suppliers = await db.suppliers.find({
            "connection_type": {"$ne": "url"},
            "ftp_host": {"$nin": [None, ""]}
        }).to_list(1000)
        url_suppliers = await db.suppliers.find({
            "connection_type": "url",
            "file_url": {"$nin": [None, ""]}
        }).to_list(1000)
        all_suppliers = ftp_suppliers + url_suppliers
        logger.info(f"Found {len(all_suppliers)} suppliers to sync")
        for supplier in all_suppliers:
            if supplier.get('ftp_paths'):
                await sync_supplier_multifile(supplier, sync_type="scheduled")
            else:
                await sync_supplier(supplier, sync_type="scheduled")
            await asyncio.sleep(2)
        logger.info("Scheduled sync completed")


async def sync_supplier_multifile(supplier: dict, sync_type: str = "manual") -> dict:
    """Sincroniza un proveedor con multiples rutas de archivo - descarga todos y fusiona por clave"""
    from services.sync.ftp_browser import resolve_latest_file

    ftp_paths = supplier.get('ftp_paths', [])
    if not ftp_paths:
        return await sync_supplier(supplier, sync_type)

    start_time = datetime.now(UTC)
    logger.info(f"Multi-file sync for {supplier['name']}: {len(ftp_paths)} files configured")
    all_file_data = {}
    all_detected_columns = {}

    _sup_hdr_raw = supplier.get('csv_header_row', 1)
    supplier_hdr = 1 if _sup_hdr_raw is None else int(_sup_hdr_raw)
    _sup_sep = supplier.get('csv_separator') or ';'
    supplier_sep = '\t' if _sup_sep == '\\t' else _sup_sep
    strip_ean_quotes = supplier.get('strip_ean_quotes', False)

    for file_config in ftp_paths:
        file_path = await resolve_latest_file(supplier, file_config)
        role = file_config.get('role', 'products')
        sep = supplier_sep
        label = file_config.get('label', file_path)

        if not file_path:
            continue

        try:
            logger.info(f"  Downloading: {file_path} (role: {role})")
            content = await download_file_from_ftp({**supplier, 'ftp_path': file_path})

            if file_path.lower().endswith('.zip') or (len(content) >= 2 and content[:2] == b'PK'):
                extracted = extract_zip_files(content)
                logger.info(f"  ZIP contains {len(extracted)} files: {list(extracted.keys())}, header_row={supplier_hdr}")
                for fname, fcontent in extracted.items():
                    rows = parse_text_file(fcontent, sep, supplier_hdr)
                    sub_role = role
                    fname_lower = fname.lower()
                    if 'stock' in fname_lower:
                        sub_role = 'stock'
                    elif 'price' in fname_lower and 'qb' not in fname_lower:
                        sub_role = 'prices'
                    elif 'qb' in fname_lower:
                        sub_role = 'prices_qb'
                    elif 'product' in fname_lower:
                        sub_role = 'products'
                    elif 'kit' in fname_lower:
                        sub_role = 'kit'
                    elif 'minqty' in fname_lower:
                        sub_role = 'min_qty'

                    if sub_role not in all_file_data or len(rows) > len(all_file_data[sub_role]):
                        all_file_data[sub_role] = rows
                        if rows:
                            all_detected_columns[sub_role] = list(rows[0].keys())
                    logger.info(f"    {fname}: {len(rows)} rows, role={sub_role}")
            else:
                rows = parse_text_file(content, sep, supplier_hdr)
                all_file_data[role] = rows
                if rows:
                    all_detected_columns[role] = list(rows[0].keys())
                logger.info(f"  {label}: {len(rows)} rows")

        except Exception as e:
            logger.error(f"  Error downloading {file_path}: {e}")
            continue

    if not all_file_data:
        return {"status": "error", "message": "No se pudo descargar ningun archivo"}

    products_data = all_file_data.get('products', [])
    prices_data = all_file_data.get('prices', [])
    stock_data = all_file_data.get('stock', [])

    if not products_data and prices_data:
        products_data = prices_data

    if not products_data:
        return {"status": "error", "message": "No se encontraron datos de productos"}

    first_row = products_data[0] if products_data else {}
    product_keys = list(first_row.keys())
    merge_key = product_keys[0] if product_keys else None

    multifile_header_row = supplier_hdr

    prices_lookup = {}
    price_merge_key = None
    if prices_data:
        price_keys = list(prices_data[0].keys()) if prices_data else []
        price_merge_key = price_keys[0] if price_keys else None
        if price_merge_key:
            for row in prices_data:
                key = str(row.get(price_merge_key, '')).strip()
                if key:
                    prices_lookup[key] = row

    stock_lookup = {}
    stock_merge_key = None
    if stock_data:
        stock_keys = list(stock_data[0].keys()) if stock_data else []
        stock_merge_key = stock_keys[0] if stock_keys else None
        if stock_merge_key:
            for row in stock_data:
                key = str(row.get(stock_merge_key, '')).strip()
                if key:
                    stock_lookup[key] = row

    logger.info(f"Merging: {len(products_data)} products, {len(prices_lookup)} prices, {len(stock_lookup)} stock entries")

    now = datetime.now(UTC).isoformat()
    errors = 0

    normalized_products = []
    column_mapping = supplier.get('column_mapping')
    for raw_product in products_data:
        try:
            prod_id = str(raw_product.get(merge_key, '')).strip()
            if not prod_id:
                errors += 1
                continue

            merged = dict(raw_product)
            if prod_id in prices_lookup:
                for k, v in prices_lookup[prod_id].items():
                    if multifile_header_row == 0:
                        if k != price_merge_key:
                            merged[f"prices_{k}"] = v
                    elif k not in merged or not merged[k]:
                        merged[k] = v
            if prod_id in stock_lookup:
                for k, v in stock_lookup[prod_id].items():
                    if multifile_header_row == 0:
                        if k != stock_merge_key:
                            merged[f"stock_{k}"] = v
                    elif k not in merged or not merged[k]:
                        merged[k] = v

            if column_mapping:
                normalized = apply_column_mapping(merged, column_mapping, strip_ean_quotes)
            else:
                normalized = normalize_product_data(merged, strip_ean_quotes)

            sku = normalized.get('sku') or prod_id
            name = normalized.get('name', '')
            if not name:
                errors += 1
                continue
            normalized_products.append(normalized)

        except Exception as e:
            logger.error(f"Error processing multi-file product: {e}")
            errors += 1

    sku_cache = await prefetch_existing_products(supplier['id'], supplier['user_id'])
    bulk_result = await bulk_upsert_products(supplier, normalized_products, sku_cache, now)
    imported = bulk_result["imported"]
    updated = bulk_result["updated"]
    errors += bulk_result["errors"]

    duration = (datetime.now(UTC) - start_time).total_seconds()
    product_count = await db.products.count_documents({"supplier_id": supplier['id']})

    flat_detected = []
    prod_cols = all_detected_columns.get('products', [])
    flat_detected.extend(prod_cols)
    price_cols = all_detected_columns.get('prices', [])
    if price_cols and multifile_header_row == 0:
        price_key = price_cols[0] if price_cols else None
        flat_detected.extend([f"prices_{c}" for c in price_cols if c != price_key])
    elif price_cols:
        flat_detected.extend([c for c in price_cols if c not in flat_detected])
    stock_cols = all_detected_columns.get('stock', [])
    if stock_cols and multifile_header_row == 0:
        stock_key = stock_cols[0] if stock_cols else None
        flat_detected.extend([f"stock_{c}" for c in stock_cols if c != stock_key])
    elif stock_cols:
        flat_detected.extend([c for c in stock_cols if c not in flat_detected])

    await db.suppliers.update_one({"id": supplier['id']}, {"$set": {
        "product_count": product_count, "last_sync": now,
        "detected_columns": flat_detected if flat_detected else list(all_detected_columns.keys())
    }})

    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), "type": "sync_complete",
        "message": f"Sincronizacion multi-archivo: {supplier['name']} - {imported} nuevos, {updated} actualizados ({len(ftp_paths)} archivos)",
        "product_id": None, "product_name": None,
        "user_id": supplier["user_id"], "read": False, "created_at": now
    })

    final_result = {
        "status": "success", "imported": imported, "updated": updated,
        "errors": errors, "files_processed": len(all_file_data),
        "detected_columns": flat_detected
    }
    await record_sync_history(supplier, final_result, sync_type, duration)

    logger.info(f"Multi-file sync complete: {imported} imported, {updated} updated, {errors} errors")
    return final_result
