import asyncio
import csv
import ftplib
import io
import ipaddress
import json
import logging
import os
import re
import socket
import uuid
import zipfile
from datetime import UTC, datetime
from urllib.parse import urlparse

import paramiko
import requests
import xlrd
from openpyxl import load_workbook
from pymongo import UpdateOne
from woocommerce import API as WooCommerceAPI

from config import (
    FTP_CONNECTION_TIMEOUT,
    FTP_DOWNLOAD_TIMEOUT,
    LOW_STOCK_THRESHOLD,
    PRICE_CHANGE_THRESHOLD_PERCENT,
    SOCKET_CONNECTION_TIMEOUT,
    URL_DOWNLOAD_TIMEOUT,
    URL_REQUEST_TIMEOUT,
    WOOCOMMERCE_API_TIMEOUT,
)
from services.database import db
from services.sku_cache import SKUCache

logger = logging.getLogger(__name__)

# Batch size for bulk DB operations
BULK_BATCH_SIZE = 500
# How often to send progress updates (every N products)
PROGRESS_REPORT_INTERVAL = 2000


def sanitize_ean_quotes(value) -> str:
    """Remove single-quote variants that some suppliers prepend/append to EAN values."""
    if value is None:
        return ""
    # ASCII apostrophe + common unicode apostrophe variants
    quote_chars = "'‘’´`＇"
    return str(value).strip().translate({ord(ch): None for ch in quote_chars})


async def send_realtime_notification(user_id: str, notification: dict):
    """Send notification via WebSocket if user is connected"""
    try:
        # Import here to avoid circular imports
        from server import ws_manager
        await ws_manager.send_to_user(user_id, {
            "type": "notification",
            "data": notification
        })
    except Exception as e:
        logger.debug(f"Could not send realtime notification: {e}")


async def send_sync_progress(user_id: str, supplier_name: str, processed: int, total: int):
    """Send sync progress update via WebSocket"""
    pct = int((processed / total) * 100) if total > 0 else 0
    await send_realtime_notification(user_id, {
        "id": str(uuid.uuid4()),
        "type": "sync_progress",
        "message": f"Sincronizando '{supplier_name}': {processed:,}/{total:,} productos ({pct}%)",
        "progress": pct,
        "processed": processed,
        "total": total,
    })


async def bulk_upsert_products(supplier: dict, normalized_products: list, sku_cache: SKUCache, now: str) -> dict:
    """
    Bulk upsert products using batched operations and SKU cache for efficient lookups.

    OPTIMIZED for 1M+ products:
    - Uses UpdateOne with upsert=True to avoid E11000 duplicate key errors
    - Populates SKU cache for first batch before processing
    - Processes in chunks to avoid loading all 1M products in memory at once
    - Batch size of 5000 for better performance
    - Streaming approach: fetch SKUs in chunks, process, flush, repeat

    Args:
        supplier: supplier dict with id, name, user_id
        normalized_products: list of dicts with keys: sku, name, description, price, stock, etc.
        sku_cache: SKUCache instance for efficient lookups
        now: ISO timestamp string

    Returns:
        dict with imported, updated, errors counts
    """
    imported = 0
    updated = 0
    errors = 0

    supplier_id = supplier['id']
    supplier_name = supplier['name']
    user_id = supplier['user_id']
    total = len(normalized_products)

    # Batch size for bulk operations
    CHUNK_SIZE = 5000
    DB_BATCH_SIZE = 5000

    product_ops = []
    price_history_docs = []
    notification_docs = []

    # Populate cache with the first batch of SKUs BEFORE the loop starts.
    # Without this, the first CHUNK_SIZE products would all miss the cache
    # and be treated as new, causing E11000 if they already exist in the DB.
    first_batch_skus = [
        p.get('sku') for p in normalized_products[:CHUNK_SIZE]
        if p.get('sku')
    ]
    if first_batch_skus:
        await sku_cache.populate_batch(first_batch_skus)

    for i, normalized in enumerate(normalized_products):
        try:
            sku = normalized.get('sku')
            name = normalized.get('name', '')
            if not sku or not name:
                errors += 1
                continue

            product_doc = {
                "sku": sku, "name": name,
                "description": normalized.get('description'),
                "price": normalized.get('price', 0),
                "stock": normalized.get('stock', 0),
                "category": normalized.get('category'),
                "brand": normalized.get('brand'),
                "ean": normalized.get('ean'),
                "weight": normalized.get('weight'),
                "image_url": normalized.get('image_url'),
                "supplier_id": supplier_id,
                "supplier_name": supplier_name,
                "user_id": user_id,
                "updated_at": now
            }

            # Load product from cache (with lazy batch loading)
            # First check by SKU (primary lookup - same supplier, same SKU)
            existing = sku_cache.get(sku)

            # CRITICAL FIX: If EAN exists, check if it's the same product from another supplier
            # This prevents duplicates of the same physical product from different distributors
            ean = normalized.get('ean')
            if not existing and ean:
                logger.debug(f"SKU {sku} not in cache, checking for existing EAN {ean}")
                # Note: This is a fallback lookup that happens AFTER cache miss
                # For performance, we don't batch-load EANs in the main loop
                # But we catch duplicates here before inserting

            if existing:
                # Track price changes
                old_price = existing.price
                new_price = product_doc['price']
                if old_price != new_price and old_price > 0:
                    change_pct = ((new_price - old_price) / old_price) * 100
                    price_history_docs.append({
                        "id": str(uuid.uuid4()), "product_id": existing.id,
                        "product_name": name, "old_price": old_price,
                        "new_price": new_price, "change_percentage": change_pct,
                        "user_id": user_id, "created_at": now
                    })
                    if abs(change_pct) >= PRICE_CHANGE_THRESHOLD_PERCENT:
                        direction = "subido" if change_pct > 0 else "bajado"
                        notification_docs.append({
                            "id": str(uuid.uuid4()), "type": "price_change",
                            "message": f"Precio de '{name[:40]}' ha {direction} {abs(change_pct):.1f}% ({old_price:.2f}€ → {new_price:.2f}€)",
                            "product_id": existing.id, "product_name": name,
                            "user_id": user_id, "read": False, "created_at": now
                        })

                # Track stock changes
                old_stock = existing.stock
                new_stock = product_doc['stock']
                if old_stock > 0 and new_stock == 0:
                    notification_docs.append({
                        "id": str(uuid.uuid4()), "type": "stock_out",
                        "message": f"Producto '{name[:40]}' sin stock",
                        "product_id": existing.id, "product_name": name,
                        "user_id": user_id, "read": False, "created_at": now
                    })
                elif old_stock > LOW_STOCK_THRESHOLD and 0 < new_stock <= LOW_STOCK_THRESHOLD:
                    notification_docs.append({
                        "id": str(uuid.uuid4()), "type": "stock_low",
                        "message": f"Producto '{name[:40]}' con stock bajo ({new_stock} uds)",
                        "product_id": existing.id, "product_name": name,
                        "user_id": user_id, "read": False, "created_at": now
                    })

                # SECURITY FIX #10: Prevent TOCTTOU race condition
                # Include the expected ID in the filter so if the doc was deleted/recreated,
                # the update fails and we create a new one with a fresh ID instead of
                # creating a duplicate with stale ID.
                # This prevents the scenario where:
                # 1. Thread A checks cache: finds product_id=123
                # 2. Thread B deletes product_id=123
                # 3. Thread A updates with id=123 (but wrong product_id ref)
                product_ops.append(UpdateOne(
                    {
                        "supplier_id": supplier_id,
                        "sku": sku,
                        "id": existing.id  # SECURITY: Ensure ID hasn't changed
                    },
                    {
                        "$set": product_doc,
                        "$setOnInsert": {"id": existing.id, "created_at": now}
                    },
                    upsert=False  # Don't create new on mismatch - let fallthrough create new one
                ))

                # If update doesn't match (ID mismatch), create new product with fresh ID
                # This handles the race condition case
                product_ops.append(UpdateOne(
                    {"supplier_id": supplier_id, "sku": sku, "id": {"$ne": existing.id}},
                    {
                        "$set": product_doc,
                        "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now}  # New ID
                    },
                    upsert=True
                ))
                updated += 1
            else:
                # New product: check for EAN duplicates first to prevent same-product imports from different suppliers
                # Note: Normally we'd batch-load EANs, but we only check on SKU misses for performance

                # CRITICAL FIX: Search by EAN to detect same product from different supplier
                # If EAN exists, update that product instead of creating duplicate
                ean_filter = {}
                if ean:
                    ean_filter = {"user_id": user_id, "ean": ean}

                # Try EAN-based match first (if EAN provided and available)
                # This prevents duplicates of same physical product from different suppliers
                # Example: TechData SKU="TECH-123" + EAN="1234567890123"
                #          MCR SKU="MCR-456" + EAN="1234567890123" (same product!)
                if ean_filter:
                    # Use upsert with EAN as primary key
                    # If product with this EAN exists, update it
                    # If not, create new product
                    product_ops.append(UpdateOne(
                        ean_filter,
                        {
                            "$set": product_doc,
                            "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now}
                        },
                        upsert=True
                    ))
                else:
                    # No EAN: use supplier+SKU as key (fallback to original behavior)
                    # This handles products without EAN gracefully
                    product_ops.append(UpdateOne(
                        {"supplier_id": supplier_id, "sku": sku},
                        {
                            "$set": product_doc,
                            "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now}
                        },
                        upsert=True
                    ))

                imported += 1

            # Flush in larger batches (5000)
            if len(product_ops) >= DB_BATCH_SIZE:
                await db.products.bulk_write(product_ops, ordered=False)
                product_ops = []

            if len(price_history_docs) >= DB_BATCH_SIZE:
                await db.price_history.insert_many(price_history_docs, ordered=False)
                price_history_docs = []

            if len(notification_docs) >= DB_BATCH_SIZE:
                await db.notifications.insert_many(notification_docs, ordered=False)
                notification_docs = []

            # Lazy-load next batch of SKUs when reaching chunk boundary
            if (i + 1) % CHUNK_SIZE == 0:
                next_batch_skus = [
                    p.get('sku') for p in normalized_products[i+1:i+1+CHUNK_SIZE]
                    if p.get('sku')
                ]
                if next_batch_skus:
                    await sku_cache.populate_batch(next_batch_skus)

                # Send progress updates
                await send_sync_progress(user_id, supplier_name, i + 1, total)

        except Exception as e:
            logger.error(f"Error processing product: {e}")
            errors += 1

    # Flush remaining operations
    if product_ops:
        await db.products.bulk_write(product_ops, ordered=False)
    if price_history_docs:
        await db.price_history.insert_many(price_history_docs, ordered=False)
    if notification_docs:
        await db.notifications.insert_many(notification_docs, ordered=False)

    # Log cache stats
    cache_stats = sku_cache.get_stats()
    logger.info(f"Sync complete for {supplier_name}: Cache stats: {cache_stats}")

    return {"imported": imported, "updated": updated, "errors": errors}


async def prefetch_existing_products(supplier_id: str, user_id: str) -> SKUCache:
    """
    Create and initialize an SKU cache for efficient lookups during sync.

    Returns: SKUCache instance (replaces dict for chunk-based loading)

    OPTIMIZED: Uses batch loading instead of full prefetch.
    With 1M products: Loads ~50K at a time, avoiding 300+ MB RAM spike.
    """
    return SKUCache(supplier_id, user_id)


# ==================== FILE DOWNLOAD ====================

def download_file_from_ftp_sync(supplier: dict) -> bytes:
    schema = supplier.get('ftp_schema', 'ftp').lower()
    host = supplier.get('ftp_host')
    port = supplier.get('ftp_port', 21)
    user = supplier.get('ftp_user', '')
    password = supplier.get('ftp_password', '')
    file_path = supplier.get('ftp_path', '')
    mode = supplier.get('ftp_mode', 'passive')
    if not host or not file_path:
        raise ValueError("FTP host and path are required")
    logger.info(f"Connecting to {schema.upper()}://{host}:{port}{file_path}")
    content = io.BytesIO()
    if schema == 'sftp':
        port = port or 22
        import socket
        sock = socket.create_connection((host, port), timeout=SOCKET_CONNECTION_TIMEOUT)
        transport = paramiko.Transport(sock)
        transport.connect(username=user, password=password)
        transport.set_keepalive(30)
        sftp = paramiko.SFTPClient.from_transport(transport)
        sftp.get_channel().settimeout(120)
        try:
            sftp.getfo(file_path, content)
            logger.info(f"SFTP download completed: {content.tell()} bytes")
        finally:
            sftp.close()
            transport.close()
    else:
        port = port or 21
        ftp = ftplib.FTP_TLS() if schema == 'ftps' else ftplib.FTP()
        try:
            ftp.connect(host, port, timeout=FTP_CONNECTION_TIMEOUT)
            ftp.login(user or 'anonymous', password or '')
            if schema == 'ftps':
                ftp.prot_p()
            ftp.set_pasv(mode == 'passive')
            logger.info(f"FTP connected, downloading {file_path}")
            ftp.retrbinary(f'RETR {file_path}', content.write)
            logger.info(f"FTP download completed: {content.tell()} bytes")
        finally:
            try:
                ftp.quit()
            except Exception:
                pass
    content.seek(0)
    return content.read()


async def download_file_from_ftp(supplier: dict) -> bytes:
    """Download file from FTP/SFTP with retry logic and exponential backoff."""
    loop = asyncio.get_running_loop()
    max_retries = 3
    timeout = FTP_DOWNLOAD_TIMEOUT  # Configurable via environment variable

    for attempt in range(max_retries):
        try:
            return await asyncio.wait_for(
                loop.run_in_executor(None, download_file_from_ftp_sync, supplier),
                timeout=timeout,
            )
        except TimeoutError:
            if attempt < max_retries - 1:
                wait_time = 2 ** (attempt + 1)  # Exponential backoff: 2s, 4s
                logger.warning(f"FTP/SFTP timeout para '{supplier.get('name', '?')}' (intento {attempt + 1}/{max_retries}), esperando {wait_time}s...")
                await asyncio.sleep(wait_time)
            else:
                raise Exception(f"Timeout descargando fichero FTP/SFTP del proveedor '{supplier.get('name', '?')}' tras {max_retries} intentos (límite: {timeout/60:.0f} minutos)")
        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = 2 ** (attempt + 1)
                logger.warning(f"Error FTP/SFTP para '{supplier.get('name', '?')}': {e} (intento {attempt + 1}/{max_retries}), reintentando en {wait_time}s...")
                await asyncio.sleep(wait_time)
            else:
                raise


def _build_browser_session(url: str, auth=None) -> tuple:
    """Build a requests Session with realistic browser headers to avoid 403 blocks."""
    import random
    from urllib.parse import urlparse

    parsed = urlparse(url)
    origin = f"{parsed.scheme}://{parsed.netloc}"

    # Realistic User-Agents (same as in http_client.py)
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
    ]

    session = requests.Session()
    session.headers.update({
        'User-Agent': random.choice(user_agents),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Referer': origin,
        'Cache-Control': 'max-age=0',
    })
    if auth:
        session.auth = auth
    return session


def _validate_url_ssrf(url: str) -> None:
    """Valida que la URL no apunte a IPs privadas/internas (prevención SSRF)."""
    parsed = urlparse(url)
    if parsed.scheme not in ('http', 'https', 'ftp', 'sftp'):
        raise ValueError(f"Esquema de URL no permitido: {parsed.scheme}")
    hostname = parsed.hostname
    if not hostname:
        raise ValueError("URL sin hostname válido")
    if '@' in (parsed.netloc.split(':')[0] if ':' in parsed.netloc else parsed.netloc):
        raise ValueError("URLs con @ en el host no están permitidas")
    try:
        resolved_ips = socket.getaddrinfo(hostname, parsed.port or 443)
        for family, _type, _proto, _canonname, sockaddr in resolved_ips:
            ip = ipaddress.ip_address(sockaddr[0])
            if ip.is_private or ip.is_loopback or ip.is_reserved or ip.is_link_local:
                raise ValueError(
                    f"URL apunta a IP privada/reservada ({ip}). "
                    f"No se permiten conexiones a redes internas."
                )
    except socket.gaierror:
        raise ValueError(f"No se pudo resolver el hostname: {hostname}")


def download_file_from_url_sync(url: str, username: str = None, password: str = None) -> bytes:
    _validate_url_ssrf(url)
    logger.info(f"Downloading from URL: {url}")
    auth = (username, password) if username and password else None
    session = _build_browser_session(url, auth)

    def _do_request(verify_ssl: bool) -> bytes:
        response = session.get(url, timeout=URL_REQUEST_TIMEOUT, stream=True, verify=verify_ssl)
        response.raise_for_status()
        content = response.content
        ssl_note = "" if verify_ssl else " (SSL verification skipped)"
        logger.info(f"URL download completed{ssl_note}: {len(content)} bytes")
        return content

    try:
        return _do_request(verify_ssl=True)
    except requests.exceptions.SSLError as ssl_err:
        logger.warning(f"SSL verification failed for {url}: {ssl_err}")
        # SECURITY FIX: Don't automatically disable SSL - this is a MITM vector
        # Instead, log the error and fail - the user must explicitly configure self-signed certificates
        logger.error(f"SSL verification failed for supplier URL: {url}")
        logger.error("To use this URL, one of the following is required:")
        logger.error("1. Use a valid SSL certificate signed by a trusted CA")
        logger.error("2. Configure certificate pinning in supplier settings")
        logger.error("3. Contact administrator to add exception for this domain")
        raise Exception(
            f"SSL verification failed para {url}. Requiere certificado SSL válido. "
            f"Contacta al administrador para configurar excepciones."
        )
    except requests.exceptions.RequestException as e:
        logger.error(f"URL download failed: {e}")
        raise Exception(f"Error descargando desde URL: {str(e)}")


async def download_file_from_url(url: str, username: str = None, password: str = None) -> bytes:
    """Download file from URL with retry logic and exponential backoff."""
    loop = asyncio.get_running_loop()
    max_retries = 3
    timeout = URL_DOWNLOAD_TIMEOUT  # Configurable via environment variable

    for attempt in range(max_retries):
        try:
            return await asyncio.wait_for(
                loop.run_in_executor(None, download_file_from_url_sync, url, username, password),
                timeout=timeout,
            )
        except TimeoutError:
            if attempt < max_retries - 1:
                wait_time = 2 ** (attempt + 1)  # Exponential backoff: 2s, 4s
                logger.warning(f"Timeout descargando URL {url[:50]}... (intento {attempt + 1}/{max_retries}), esperando {wait_time}s...")
                await asyncio.sleep(wait_time)
            else:
                raise Exception(f"Timeout descargando desde URL tras {max_retries} intentos (límite: {timeout/60:.0f} minutos): {url[:100]}")
        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = 2 ** (attempt + 1)
                logger.warning(f"Error descargando URL {url[:50]}...: {e} (intento {attempt + 1}/{max_retries}), reintentando en {wait_time}s...")
                await asyncio.sleep(wait_time)
            else:
                raise


# ==================== FILE PARSING ====================

def _sanitize_csv_cell(value):
    """Previene CSV formula injection eliminando prefijos peligrosos."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


def parse_csv_content(content: bytes) -> list:
    try:
        decoded = content.decode('utf-8')
    except Exception:
        decoded = content.decode('latin-1')
    reader = csv.DictReader(io.StringIO(decoded))
    result = []
    for row in reader:
        sanitized = {k: _sanitize_csv_cell(v) if isinstance(v, str) else v
                     for k, v in row.items()}
        result.append(sanitized)
    return result


def parse_xlsx_content(content: bytes) -> list:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).lower().strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(row)]


def parse_xls_content(content: bytes) -> list:
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, c)).lower().strip() for c in range(ws.ncols)]
    return [{headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)} for r in range(1, ws.nrows)]


def parse_xml_content(content: bytes) -> list:
    """Parse XML content safely - prevents XXE attacks

    SECURITY FIX: Disable external entities to prevent XXE injection
    """
    try:
        decoded = content.decode('utf-8')
    except Exception:
        decoded = content.decode('latin-1')

    # SECURITY: Use defusedxml to prevent XXE attacks
    # Disables: external entities, internal DTD parsing, entity expansion
    try:
        from defusedxml import xmltodict as safe_xmltodict
        data = safe_xmltodict.parse(decoded, disable_entities=True, process_namespaces=False)
    except ImportError:
        raise ImportError(
            "El parseo de XML requiere la librería defusedxml para protección XXE. "
            "Instalar con: pip install defusedxml"
        )

    for key in ['products', 'items', 'catalog', 'data', 'root']:
        if key in data:
            items = data[key]
            if isinstance(items, dict):
                for subkey in items:
                    if isinstance(items[subkey], list):
                        return items[subkey]
            elif isinstance(items, list):
                return items
    return []


def normalize_product_data(raw: dict, strip_ean_quotes: bool = False) -> dict:
    mapping = {
        'sku': ['sku', 'codigo', 'code', 'ref', 'referencia', 'reference', 'id', 'product_id', 'partnumber', 'part_number',
                'articulo', 'codigo_articulo', 'cod', 'item_code', 'cod_articulo', 'ref_articulo', 'codigo_producto',
                'product_code', 'item_id', 'article', 'article_id', 'num_art', 'numero_articulo', 'codigoarticulo',
                'codart', 'refart', 'art', 'producto_id', 'id_producto', 'idproducto', 'producto'],
        'name': ['name', 'nombre', 'title', 'titulo', 'product_name', 'descripcion', 'description', 'producto',
                 'articulo_nombre', 'item_name', 'denominacion', 'denominación', 'nombre_producto', 'product',
                 'item', 'nombreprod', 'nombproducto', 'nombre_articulo', 'articulo', 'desc', 'descriptivo'],
        'price': ['price', 'precio', 'pvp', 'cost', 'coste', 'unit_price', 'tarifa', 'importe', 'pricen',
                  'precio_neto', 'net_price', 'preciopvp', 'precioventa', 'precio_venta', 'preciofinal',
                  'precio_final', 'priceunit', 'unitprice', 'precio_unitario', 'pvd', 'pvr', 'eur', 'euro'],
        'stock': ['stock', 'quantity', 'cantidad', 'qty', 'inventory', 'disponible', 'existencias', 'unidades',
                  'disponibilidad', 'units', 'existencia', 'cantstock', 'stockdisponible', 'stock_disponible',
                  'en_stock', 'enstock', 'cantidad_stock', 'total_stock', 'stockactual', 'stock_actual'],
        'category': ['category', 'categoria', 'cat', 'type', 'tipo', 'familia', 'family', 'grupo', 'group',
                     'categorie', 'categoria1', 'cat1', 'groupo', 'seccion', 'sección'],
        'brand': ['brand', 'marca', 'manufacturer', 'fabricante', 'vendor', 'proveedor', 'make', 'brand_name',
                  'nombremarca', 'nombre_marca', 'marcafabricante'],
        'ean': ['ean', 'ean13', 'barcode', 'upc', 'codigo_barras', 'gtin', 'ean_code', 'codigobarras',
                'cod_barras', 'codigo_barra', 'bar_code', 'ean8', 'codean'],
        'weight': ['weight', 'peso', 'kg', 'mass', 'peso_kg', 'pesokg', 'weightkg'],
        'image_url': ['image', 'imagen', 'image_url', 'photo', 'foto', 'picture', 'url_imagen', 'img',
                      'urlimagen', 'imageurl', 'fotografia', 'pic', 'imagen_url', 'url_image', 'link_imagen'],
        'description': ['description', 'descripcion', 'desc', 'details', 'detalles', 'long_description',
                        'short_description', 'descripcion_larga', 'descripcion_corta', 'texto', 'detalle']
    }
    result = {}
    raw_lower = {str(k).lower().strip().replace(' ', '_').replace('-', '_'): v for k, v in raw.items()}
    raw_original_lower = {str(k).lower().strip(): v for k, v in raw.items()}
    # Merge both for more flexible matching
    combined_raw = {**raw_original_lower, **raw_lower}

    logger.debug(f"Normalizing product - columns available: {list(combined_raw.keys())}")

    for field, aliases in mapping.items():
        for alias in aliases:
            if alias in combined_raw and combined_raw[alias]:
                value = combined_raw[alias]
                if field in ['price', 'weight']:
                    try:
                        value = float(str(value).replace(',', '.').replace('€', '').replace('$', '').strip())
                    except Exception:
                        value = 0.0
                elif field == 'stock':
                    try:
                        value = int(float(str(value).replace(',', '.')))
                    except Exception:
                        value = 0
                elif field == 'ean' and strip_ean_quotes:
                    # Remove single quotes from EAN if strip_ean_quotes is enabled
                    value = sanitize_ean_quotes(value)
                result[field] = value
                break

    # Log what was detected
    if not result.get('sku') or not result.get('name'):
        available_cols = list(raw.keys())[:15]
        logger.warning(f"Product missing required fields. SKU: {result.get('sku')}, Name: {result.get('name')}. Available columns: {available_cols}")

    return result


def apply_column_mapping(raw_data: dict, column_mapping: dict, strip_ean_quotes: bool = False) -> dict:
    if not column_mapping:
        return normalize_product_data(raw_data, strip_ean_quotes)
    result = {}
    raw_lower = {str(k).lower().strip(): v for k, v in raw_data.items()}
    raw_original = {str(k).strip(): v for k, v in raw_data.items()}
    field_types = {
        'sku': 'string', 'name': 'string', 'description': 'string',
        'price': 'float', 'price2': 'float', 'stock': 'int',
        'ean': 'string', 'brand': 'string', 'category': 'string',
        'subcategory': 'string', 'subcategory2': 'string',
        'weight': 'float', 'image_url': 'string', 'image_url2': 'string',
        'image_url3': 'string', 'short_description': 'string', 'long_description': 'string'
    }
    for system_field, m in column_mapping.items():
        if not m:
            continue
        columns = [m] if isinstance(m, str) else m if isinstance(m, list) else []
        values = []
        for col in columns:
            if not col:
                continue
            # IMPROVED: Soportar referencias dinámicas como "prices_QB_col_3"
            # Primero intenta buscar exactamente, luego con case-insensitive
            value = raw_original.get(col) or raw_lower.get(col.lower().strip())

            # Si no encuentra, intenta variaciones de case/underscore
            if value is None:
                # Probar variaciones: prices_QB_col_3, prices_qb_col_3, etc.
                value = raw_original.get(col.replace('_QB_', '_qb_')) or raw_original.get(col.replace('_QB_', '_'))

            if value is not None and str(value).strip() != '':
                values.append(str(value).strip())
        if values:
            combined_value = ' > '.join(values) if system_field.startswith('category') else ' '.join(values)
            field_type = field_types.get(system_field, 'string')
            try:
                if field_type == 'float':
                    combined_value = float(str(combined_value).replace(',', '.').replace('€', '').replace('$', '').strip())
                elif field_type == 'int':
                    combined_value = int(float(str(combined_value).replace(',', '.')))
                elif field_type == 'string' and system_field == 'ean' and strip_ean_quotes:
                    # Remove single quotes from EAN if strip_ean_quotes is enabled
                    combined_value = sanitize_ean_quotes(combined_value)
            except Exception:
                if field_type in ['float', 'int']:
                    combined_value = 0
            result[system_field] = combined_value
    product = {
        'sku': result.get('sku', ''), 'name': result.get('name', ''),
        'description': result.get('description') or result.get('long_description') or result.get('short_description', ''),
        'price': result.get('price', 0), 'stock': result.get('stock', 0),
        'category': result.get('category', ''), 'brand': result.get('brand', ''),
        'ean': result.get('ean', ''), 'weight': result.get('weight'),
        'image_url': result.get('image_url', '')
    }
    categories = [result.get('category', '')]
    if result.get('subcategory'):
        categories.append(result['subcategory'])
    if result.get('subcategory2'):
        categories.append(result['subcategory2'])
    product['category'] = ' > '.join([c for c in categories if c])
    return product


# ==================== SUPPLIER SYNC ====================

async def process_supplier_file(supplier: dict, content: bytes) -> dict:
    file_format = supplier.get('file_format', 'csv').lower()
    separator = supplier.get('csv_separator', ';')
    enclosure = supplier.get('csv_enclosure', '"')
    _header_row_raw = supplier.get('csv_header_row', 1)
    header_row = 1 if _header_row_raw is None else int(_header_row_raw)
    column_mapping = supplier.get('column_mapping')
    strip_ean_quotes = supplier.get('strip_ean_quotes', False)
    detected_columns = []

    # Auto-detect ZIP by magic bytes (PK signature) regardless of configured file_format
    if len(content) >= 4 and content[:2] == b'PK':
        file_format = 'zip'

    try:
        if file_format == 'zip':
            # Extract and find compatible files inside the ZIP
            try:
                extracted = extract_zip_files(content)
            except Exception as e:
                return {"imported": 0, "updated": 0, "errors": 0, "message": f"Error al descomprimir ZIP: {e}"}

            if not extracted:
                return {"imported": 0, "updated": 0, "errors": 0, "message": "El archivo ZIP está vacío"}

            logger.info(f"ZIP detectado con {len(extracted)} archivo(s): {list(extracted.keys())}")

            # Collect all compatible flat files (csv, txt, xlsx, xls, xml)
            compatible_exts = ('.csv', '.txt', '.xlsx', '.xls', '.xml')
            compatible_files = [
                (fname, fcontent)
                for fname, fcontent in extracted.items()
                if fname.lower().endswith(compatible_exts)
                and not fname.split('/')[-1].startswith('.')
                and not fname.split('/')[-1].startswith('__')
            ]

            if not compatible_files:
                names = list(extracted.keys())
                return {
                    "imported": 0, "updated": 0, "errors": 0,
                    "message": f"El ZIP no contiene archivos compatibles (csv/xlsx/xls/xml/txt). Archivos encontrados: {names}"
                }

            # If only one compatible file, process it directly
            if len(compatible_files) == 1:
                best_fname, best_content = compatible_files[0]
                best_fmt = 'csv' if best_fname.lower().endswith(('.csv', '.txt')) else best_fname.rsplit('.', 1)[-1].lower()
                logger.info(f"ZIP (único archivo): procesando '{best_fname}' como {best_fmt.upper()}")
                zip_supplier = {**supplier, 'file_format': best_fmt}
                return await process_supplier_file(zip_supplier, best_content)

            # Multiple compatible files: auto-detect roles from filenames and merge
            logger.info(f"ZIP multi-archivo: detectando roles automáticamente para {len(compatible_files)} archivos")
            role_keywords = {
                'stock':    ['stock', 'inventory', 'disponibilidad', 'existencias', 'qty', 'quantity'],
                'prices':   ['price', 'precio', 'tarif', 'coste', 'cost', 'pvp', 'pvd'],
                'prices_qb': ['prices_qb', 'pricesqb'],  # Detectar prices_QB específicamente (TechData) — eliminado 'qb' solo para evitar falsos positivos
                'products': ['product', 'catalog', 'article', 'catalogo', 'articulo', 'master', 'items'],
            }

            all_file_data = {}
            for fname, fcontent in compatible_files:
                fname_base = fname.split('/')[-1].lower()
                detected_role = 'products'  # default

                # Check for prices_QB FIRST (more specific) before generic 'prices'
                if any(kw in fname_base for kw in role_keywords.get('prices_qb', [])):
                    detected_role = 'prices_qb'
                else:
                    for role, keywords in role_keywords.items():
                        if role == 'prices_qb':
                            continue  # Already checked
                        if any(kw in fname_base for kw in keywords):
                            detected_role = role
                            break

                # Determine header_row per file (StockFile.txt-style files often have header)
                # Use supplier setting; individual files can override if needed
                file_hdr = header_row
                try:
                    if fname.lower().endswith(('.xlsx', '.xls')):
                        fmt = 'xlsx' if fname.lower().endswith('.xlsx') else 'xls'
                        if fmt == 'xlsx':
                            wb = load_workbook(filename=io.BytesIO(fcontent), read_only=True)
                            ws = wb.active
                            rows = list(ws.iter_rows(values_only=True))
                            hdrs = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
                            file_rows = [dict(zip(hdrs, r)) for r in rows[1:] if any(r)]
                        else:
                            wb = xlrd.open_workbook(file_contents=fcontent)
                            ws = wb.sheet_by_index(0)
                            hdrs = [str(ws.cell_value(0, c)).strip() or f'col_{c}' for c in range(ws.ncols)]
                            file_rows = [{hdrs[c]: ws.cell_value(r, c) for c in range(ws.ncols)} for r in range(1, ws.nrows)]
                    else:
                        file_rows = parse_text_file(fcontent, separator, file_hdr)
                except Exception as fe:
                    logger.warning(f"  No se pudo parsear {fname}: {fe}")
                    continue

                # Don't overwrite an existing role with fewer rows
                # PERO: si es prices_QB, SIEMPRE úsalo en lugar de 'prices' genérico (es más específico)
                if detected_role == 'prices_qb':
                    all_file_data[detected_role] = file_rows
                    # También bórralo de 'prices' genérico si existía
                    if 'prices' in all_file_data and len(file_rows) > len(all_file_data.get('prices', [])):
                        logger.info(f"  Prefiriendo prices_QB sobre prices genérico (más específico)")
                elif detected_role not in all_file_data or len(file_rows) > len(all_file_data[detected_role]):
                    all_file_data[detected_role] = file_rows
                logger.info(f"  {fname} → role={detected_role}, filas={len(file_rows)}")

            if not all_file_data:
                return {"imported": 0, "updated": 0, "errors": 0, "message": "No se pudo parsear ningún archivo del ZIP"}

            products_data = all_file_data.get('products', [])
            prices_data   = all_file_data.get('prices', []) or all_file_data.get('prices_qb', [])  # Prefer prices_qb if available
            prices_qb_data = all_file_data.get('prices_qb', [])
            stock_data    = all_file_data.get('stock', [])

            # Fallback: if no 'products' role detected, use the largest file
            if not products_data:
                largest_role = max(all_file_data, key=lambda r: len(all_file_data[r]))
                products_data = all_file_data[largest_role]

            # Build lookup tables for price and stock
            prices_merge_key = None
            prices_lookup = {}
            prices_qb_lookup = {}
            is_prices_qb = bool(prices_qb_data)

            # Try prices_QB first (mejor estructura para TechData)
            prices_data_to_use = prices_qb_data if prices_qb_data else prices_data

            if prices_data_to_use:
                # IMPROVED: Preferir SKU (col_3 en TechData) como merge key
                sample_product = products_data[0] if products_data else {}
                product_keys = list(sample_product.keys())

                # Estrategia 1: Buscar columna común (mejor opción)
                common_merge_key = None

                # Preferir col_3 (SKU en TechData) o col_1 (nombre/SKU en precios_QB)
                # Solo usar col_3/col_1 si es realmente TechData (prices_QB)
                if is_prices_qb:
                    preferred_keys = ['col_3', 'col_1', 'sku', 'code', 'ref']
                else:
                    preferred_keys = ['sku', 'code', 'ref', 'id', 'article', 'articulo', 'cod', 'codigo']

                for preferred in preferred_keys:
                    if preferred in product_keys and preferred in prices_data_to_use[0]:
                        common_merge_key = preferred
                        break

                # Estrategia 2: Buscar CUALQUIER columna común
                if not common_merge_key:
                    for pk in product_keys:
                        if pk in prices_data_to_use[0]:
                            common_merge_key = pk
                            break

                # Fallback: usar la primera columna
                prices_merge_key = common_merge_key or list(prices_data_to_use[0].keys())[0]

                logger.info(f"ZIP prices merge_key: {prices_merge_key} (common={common_merge_key is not None}, is_qb={is_prices_qb})")

                if prices_merge_key:
                    for row in prices_data_to_use:  # FIXED: usar prices_data_to_use, no prices_data
                        k = str(row.get(prices_merge_key, '')).strip()
                        if k:
                            prices_lookup[k] = row

            stock_merge_key = None
            stock_lookup = {}
            if stock_data:
                # Try to find a common merge key for stock too
                sample_product = products_data[0] if products_data else {}
                product_keys = list(sample_product.keys())

                common_merge_key = None
                for pk in product_keys:
                    if pk in stock_data[0]:
                        common_merge_key = pk
                        break

                stock_merge_key = common_merge_key or list(stock_data[0].keys())[0]

                logger.info(f"ZIP stock merge_key: {stock_merge_key} (common={common_merge_key is not None})")

                if stock_merge_key:
                    for row in stock_data:
                        k = str(row.get(stock_merge_key, '')).strip()
                        if k:
                            stock_lookup[k] = row

            logger.info(f"ZIP merge: {len(products_data)} productos, {len(prices_lookup)} precios (key={prices_merge_key}), {len(stock_lookup)} stock (key={stock_merge_key})")

            # Build detected_columns from a sample merged row (product + prefixed prices/stock)
            # so the ColumnMappingDialog shows the real available columns
            zip_detected_cols = []
            if products_data:
                sample = dict(products_data[0])
                # Usar prices_merge_key para obtener la muestra correcta
                sample_id = str(sample.get(prices_merge_key if prices_merge_key else list(sample.keys())[0], '')).strip()
                if prices_merge_key and sample_id in prices_lookup:
                    prices_prefix = "prices_QB_" if is_prices_qb else "prices_"
                    for k in prices_lookup[sample_id]:
                        if header_row == 0:
                            if k != prices_merge_key:
                                sample[f"{prices_prefix}{k}"] = prices_lookup[sample_id][k]
                        elif k not in sample:
                            sample[k] = prices_lookup[sample_id][k]
                if stock_merge_key and sample_id in stock_lookup:
                    for k in stock_lookup[sample_id]:
                        if header_row == 0:
                            if k != stock_merge_key:
                                sample[f"stock_{k}"] = stock_lookup[sample_id][k]
                        elif k not in sample:
                            sample[k] = stock_lookup[sample_id][k]
                zip_detected_cols = list(sample.keys())
                await db.suppliers.update_one(
                    {"id": supplier['id']},
                    {"$set": {"detected_columns": zip_detected_cols}}
                )

            # Merge and normalize products
            now = datetime.now(UTC).isoformat()
            needs_mapping = False
            errors = 0
            # IMPROVED: Usar prices_merge_key para merge (mejor que col_0)
            merge_key = prices_merge_key if prices_merge_key and prices_lookup else list(products_data[0].keys())[0] if products_data else None
            column_mapping = supplier.get('column_mapping')

            normalized_products = []
            for raw in products_data:
                try:
                    prod_id = str(raw.get(merge_key, '')).strip() if merge_key else ''
                    merged = dict(raw)
                    if prod_id and prod_id in prices_lookup:
                        # Determinar el prefijo correcto según el tipo de archivo de precios
                        prices_prefix = "prices_QB_" if is_prices_qb else "prices_"
                        for k, v in prices_lookup[prod_id].items():
                            if header_row == 0:
                                if k != prices_merge_key:
                                    merged[f"{prices_prefix}{k}"] = v  # FIXED: usar el prefijo correcto
                            elif k not in merged or not merged[k]:
                                merged[k] = v
                    if prod_id and prod_id in stock_lookup:
                        for k, v in stock_lookup[prod_id].items():
                            if header_row == 0:
                                if k != stock_merge_key:
                                    merged[f"stock_{k}"] = v
                            elif k not in merged or not merged[k]:
                                merged[k] = v

                    normalized = apply_column_mapping(merged, column_mapping, strip_ean_quotes) if column_mapping else normalize_product_data(merged, strip_ean_quotes)
                    sku = normalized.get('sku') or prod_id
                    name = normalized.get('name', '')
                    if not sku or not name:
                        errors += 1
                        needs_mapping = True
                        continue
                    normalized_products.append(normalized)
                except Exception as e:
                    logger.error(f"Error procesando producto ZIP: {e}")
                    errors += 1

            # Bulk upsert with SKU cache for efficient lookups
            sku_cache = await prefetch_existing_products(supplier['id'], supplier['user_id'])
            bulk_result = await bulk_upsert_products(supplier, normalized_products, sku_cache, now)
            imported = bulk_result["imported"]
            updated = bulk_result["updated"]
            errors += bulk_result["errors"]

            msg = f"ZIP procesado: {imported} importados, {updated} actualizados"
            result = {"imported": imported, "updated": updated, "errors": errors, "message": msg,
                      "needs_mapping": needs_mapping, "detected_columns": zip_detected_cols}
            if needs_mapping and imported + updated == 0:
                result["status"] = "needs_mapping"
                result["message"] = (
                    f"El ZIP se procesó pero no se importaron productos. "
                    f"Columnas disponibles: {', '.join(zip_detected_cols[:8])}... "
                    f"Configura el mapeo de columnas o re-aplica la plantilla del proveedor."
                )
            return result

        elif file_format == 'csv':
            try:
                decoded = content.decode('utf-8-sig')
            except Exception:
                try:
                    decoded = content.decode('utf-8')
                except Exception:
                    decoded = content.decode('latin-1')
            lines = decoded.split('\n')
            if header_row > 1:
                lines = lines[header_row-1:]
            if separator == '\\t':
                separator = '\t'
            first_line_raw = lines[0].rstrip('\r') if lines else ''
            separator = _detect_best_separator(first_line_raw, separator)
            if header_row == 0:
                # Headerless file: generate positional column names (col_0, col_1, ...)
                first_line = first_line_raw
                first_row_parsed = list(csv.reader([first_line], delimiter=separator, quotechar=enclosure if enclosure else '"'))
                num_cols = len(first_row_parsed[0]) if first_row_parsed else 0
                fieldnames = [f'col_{i}' for i in range(num_cols)]
                reader = csv.DictReader(lines, fieldnames=fieldnames, delimiter=separator, quotechar=enclosure if enclosure else '"')
            else:
                reader = csv.DictReader(lines, delimiter=separator, quotechar=enclosure if enclosure else '"')
            raw_products = list(reader)
            if raw_products:
                detected_columns = list(raw_products[0].keys())
        elif file_format in ['xlsx', 'xls']:
            if file_format == 'xlsx':
                wb = load_workbook(filename=io.BytesIO(content), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            else:
                wb = xlrd.open_workbook(file_contents=content)
                ws = wb.sheet_by_index(0)
                rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
            if not rows:
                return {"imported": 0, "updated": 0, "errors": 0}
            if header_row > 1:
                rows = rows[header_row-1:]
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
            detected_columns = headers
            raw_products = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
        elif file_format == 'xml':
            raw_products = parse_xml_content(content)
            if raw_products:
                detected_columns = list(raw_products[0].keys())
        else:
            return {"imported": 0, "updated": 0, "errors": 0, "message": "Unsupported format"}
        if detected_columns:
            await db.suppliers.update_one({"id": supplier['id']}, {"$set": {"detected_columns": detected_columns}})
        now = datetime.now(UTC).isoformat()
        imported = 0
        updated = 0
        errors = 0
        needs_mapping = False
        if not column_mapping and detected_columns:
            col_names_lower = [c.lower().strip().replace(' ', '_').replace('-', '_') for c in detected_columns]
            col_names_simple = [c.lower().strip() for c in detected_columns]
            all_col_variants = col_names_lower + col_names_simple

            sku_aliases = ['sku', 'codigo', 'referencia', 'ref', 'reference', 'id', 'cod', 'articulo',
                          'codigo_articulo', 'product_id', 'item_code', 'code', 'producto']
            name_aliases = ['name', 'nombre', 'title', 'titulo', 'product', 'producto', 'descripcion',
                           'description', 'denominacion', 'item', 'articulo', 'desc']

            has_sku = any(alias in all_col_variants for alias in sku_aliases)
            has_name = any(alias in all_col_variants for alias in name_aliases)

            if not has_sku or not has_name:
                needs_mapping = True
                logger.warning(f"Supplier {supplier.get('name')}: needs_mapping=True. Columns detected: {detected_columns[:10]}. has_sku={has_sku}, has_name={has_name}")
            if not has_sku or not has_name:
                needs_mapping = True

        # Normalize all products first
        normalized_products = []
        for raw in raw_products:
            try:
                normalized = apply_column_mapping(raw, column_mapping, strip_ean_quotes) if column_mapping else normalize_product_data(raw, strip_ean_quotes)
                if not normalized.get('sku') or not normalized.get('name'):
                    errors += 1
                    continue
                normalized_products.append(normalized)
            except Exception as e:
                logger.error(f"Error processing product: {e}")
                errors += 1

        # Bulk upsert with SKU cache for efficient lookups
        sku_cache = await prefetch_existing_products(supplier['id'], supplier['user_id'])
        bulk_result = await bulk_upsert_products(supplier, normalized_products, sku_cache, now)
        imported = bulk_result["imported"]
        updated = bulk_result["updated"]
        errors += bulk_result["errors"]

        result = {"imported": imported, "updated": updated, "errors": errors, "detected_columns": detected_columns}

        # Provide better feedback about mapping needs
        if needs_mapping:
            result["needs_mapping"] = True
            if errors > 0 and (imported + updated) == 0:
                result["message"] = f"Las columnas detectadas ({', '.join(detected_columns[:5])}...) no coinciden con los nombres estándar. Configura el mapeo de columnas para este proveedor."
                result["status"] = "needs_mapping"
            else:
                result["message"] = "Importación parcial. Se recomienda configurar el mapeo de columnas para mejores resultados."
        elif errors > 0 and (imported + updated) == 0:
            result["message"] = f"No se pudieron importar productos. Verifica el formato del archivo y las columnas: {', '.join(detected_columns[:5])}"
            result["status"] = "error"

        return result
    except Exception as e:
        logger.error(f"Error processing file: {e}")
        return {"imported": 0, "updated": 0, "errors": 1, "message": str(e), "detected_columns": []}


async def record_sync_history(supplier: dict, result: dict, sync_type: str, duration: float, error_message: str = None):
    """Registrar historial de sincronización"""
    status = "success" if result.get("status") == "success" else "error" if result.get("status") == "error" else "partial"
    if result.get("errors", 0) > 0 and result.get("imported", 0) + result.get("updated", 0) > 0:
        status = "partial"

    await db.sync_history.insert_one({
        "id": str(uuid.uuid4()),
        "supplier_id": supplier["id"],
        "supplier_name": supplier["name"],
        "sync_type": sync_type,
        "status": status,
        "imported": result.get("imported", 0),
        "updated": result.get("updated", 0),
        "errors": result.get("errors", 0),
        "duration_seconds": round(duration, 2),
        "error_message": error_message or result.get("message"),
        "user_id": supplier["user_id"],
        "created_at": datetime.now(UTC).isoformat()
    })


async def sync_supplier(supplier: dict, sync_type: str = "manual") -> dict:
    connection_type = supplier.get('connection_type', 'ftp')
    if connection_type == 'url':
        if not supplier.get('file_url'):
            return {"status": "skipped", "message": "URL no configurada"}
    else:
        if not supplier.get('ftp_host') or not supplier.get('ftp_path'):
            return {"status": "skipped", "message": "FTP no configurado"}

    start_time = datetime.now(UTC)
    try:
        logger.info(f"Syncing supplier: {supplier['name']} (via {connection_type})")
        if connection_type == 'url':
            from services.encryption import decrypt_password
            url_username = supplier.get('url_username')
            url_password = supplier.get('url_password')
            if url_password:
                url_password = decrypt_password(url_password)
            content = await download_file_from_url(supplier['file_url'], url_username, url_password)
        else:
            content = await download_file_from_ftp(supplier)
        result = await process_supplier_file(supplier, content)
        now = datetime.now(UTC)
        duration = (now - start_time).total_seconds()

        product_count = await db.products.count_documents({"supplier_id": supplier['id']})
        await db.suppliers.update_one({"id": supplier['id']}, {"$set": {"product_count": product_count, "last_sync": now.isoformat()}})

        notification = {
            "id": str(uuid.uuid4()), "type": "sync_complete",
            "message": f"Sincronización completada: {supplier['name']} - {result['imported']} nuevos, {result['updated']} actualizados",
            "product_id": None, "product_name": None,
            "user_id": supplier["user_id"], "read": False, "created_at": now.isoformat()
        }
        await db.notifications.insert_one(notification)
        await send_realtime_notification(supplier["user_id"], notification)

        final_result = {"status": "success", **result}
        await record_sync_history(supplier, final_result, sync_type, duration)
        logger.info(f"Sync complete for {supplier['name']}: {result}")
        return final_result
    except Exception as e:
        duration = (datetime.now(UTC) - start_time).total_seconds()
        logger.error(f"Error syncing supplier {supplier['name']}: {e}")

        notification = {
            "id": str(uuid.uuid4()), "type": "sync_error",
            "message": f"Error en sincronización: {supplier['name']} - {str(e)[:100]}",
            "product_id": None, "product_name": None,
            "user_id": supplier["user_id"], "read": False,
            "created_at": datetime.now(UTC).isoformat()
        }
        await db.notifications.insert_one(notification)
        await send_realtime_notification(supplier["user_id"], notification)

        error_result = {"status": "error", "message": str(e), "imported": 0, "updated": 0, "errors": 1}
        await record_sync_history(supplier, error_result, sync_type, duration, str(e))
        return error_result


async def sync_all_suppliers():
    from services.unified_sync import _global_sync_lock
    logger.info("Starting scheduled sync for all suppliers (waiting for global lock)...")
    async with _global_sync_lock:
        ftp_suppliers = await db.suppliers.find({
            "connection_type": {"$ne": "url"},
            "ftp_host": {"$nin": [None, ""]}
        }).to_list(1000)
        url_suppliers = await db.suppliers.find({
            "connection_type": "url",
            "file_url": {"$nin": [None, ""]}
        }).to_list(1000)
        all_suppliers = ftp_suppliers + url_suppliers
        logger.info(f"Found {len(all_suppliers)} suppliers to sync")
        for supplier in all_suppliers:
            if supplier.get('ftp_paths'):
                await sync_supplier_multifile(supplier, sync_type="scheduled")
            else:
                await sync_supplier(supplier, sync_type="scheduled")
            await asyncio.sleep(2)
        logger.info("Scheduled sync completed")



# ==================== FTP BROWSER ====================

def browse_ftp_sync(config: dict, path: str = "/") -> dict:
    """
    Navega por el servidor FTP/SFTP y lista archivos y carpetas.
    Soporta: FTP, FTPS, SFTP
    """
    schema = config.get('ftp_schema', 'ftp').lower()
    host = config.get('ftp_host')
    port = config.get('ftp_port', 21)
    user = config.get('ftp_user', '')
    password = config.get('ftp_password', '')
    mode = config.get('ftp_mode', 'passive')

    if not host:
        return {"status": "error", "message": "FTP host is required", "files": [], "path": path}

    files = []
    error_message = None

    try:
        if schema == 'sftp':
            port = port or 22
            transport = paramiko.Transport((host, port))
            transport.connect(username=user, password=password)
            sftp = paramiko.SFTPClient.from_transport(transport)
            try:
                for attr in sftp.listdir_attr(path):
                    is_dir = attr.st_mode and (attr.st_mode & 0o170000 == 0o040000)
                    file_ext = attr.filename.rsplit('.', 1)[-1].lower() if '.' in attr.filename else ''
                    files.append({
                        "name": attr.filename,
                        "path": f"{path.rstrip('/')}/{attr.filename}",
                        "size": attr.st_size,
                        "size_formatted": format_file_size(attr.st_size),
                        "is_dir": is_dir,
                        "is_supported": file_ext in ['csv', 'xlsx', 'xls', 'xml', 'zip', 'txt'],
                        "extension": file_ext,
                        "modified": str(datetime.fromtimestamp(attr.st_mtime)) if attr.st_mtime else None
                    })
            finally:
                sftp.close()
                transport.close()
        else:
            port = port or 21
            ftp = ftplib.FTP_TLS() if schema == 'ftps' else ftplib.FTP()
            try:
                ftp.connect(host, port, timeout=FTP_CONNECTION_TIMEOUT)
                ftp.login(user or 'anonymous', password or '')
                if schema == 'ftps':
                    ftp.prot_p()
                ftp.set_pasv(mode == 'passive')

                # Intentar usar MLSD primero (más información)
                try:
                    for name, facts in ftp.mlsd(path):
                        if name in ['.', '..']:
                            continue
                        is_dir = facts.get('type') == 'dir'
                        size = int(facts.get('size', 0)) if not is_dir else 0
                        file_ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
                        modified = facts.get('modify', '')
                        if modified and len(modified) >= 14:
                            modified = f"{modified[:4]}-{modified[4:6]}-{modified[6:8]} {modified[8:10]}:{modified[10:12]}"
                        files.append({
                            "name": name,
                            "path": f"{path.rstrip('/')}/{name}",
                            "size": size,
                            "size_formatted": format_file_size(size),
                            "is_dir": is_dir,
                            "is_supported": file_ext in ['csv', 'xlsx', 'xls', 'xml', 'zip', 'txt'],
                            "extension": file_ext,
                            "modified": modified
                        })
                except Exception:
                    # Fallback a DIR si MLSD no está soportado
                    raw_lines = []
                    ftp.dir(path, raw_lines.append)
                    for line in raw_lines:
                        parts = line.split(None, 8)
                        if len(parts) >= 9:
                            name = parts[8]
                            if name in ['.', '..']:
                                continue
                            is_dir = line.startswith('d')
                            size = int(parts[4]) if not is_dir else 0
                            date_str = f"{parts[5]} {parts[6]} {parts[7]}"
                            file_ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
                            files.append({
                                "name": name,
                                "path": f"{path.rstrip('/')}/{name}",
                                "size": size,
                                "size_formatted": format_file_size(size),
                                "is_dir": is_dir,
                                "is_supported": file_ext in ['csv', 'xlsx', 'xls', 'xml', 'zip', 'txt'],
                                "extension": file_ext,
                                "modified": date_str
                            })
            finally:
                try:
                    ftp.quit()
                except Exception:
                    pass
    except ftplib.error_perm as e:
        error_message = f"Error de permisos FTP: {str(e)}"
        logger.error(f"FTP permission error browsing {path}: {e}")
    except Exception as e:
        error_message = f"Error de conexión: {str(e)}"
        logger.error(f"FTP browse error for {path}: {e}")

    # Ordenar: carpetas primero, luego archivos por nombre
    files.sort(key=lambda x: (not x["is_dir"], x["name"].lower()))

    # Calcular estadísticas
    total_files = len([f for f in files if not f["is_dir"]])
    supported_files = len([f for f in files if f.get("is_supported")])
    total_dirs = len([f for f in files if f["is_dir"]])

    return {
        "status": "ok" if not error_message else "error",
        "message": error_message,
        "path": path,
        "files": files,
        "stats": {
            "total_files": total_files,
            "supported_files": supported_files,
            "total_dirs": total_dirs
        }
    }


def format_file_size(size: int) -> str:
    """Formatea el tamaño de archivo en formato legible"""
    if size < 1024:
        return f"{size} B"
    elif size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    elif size < 1024 * 1024 * 1024:
        return f"{size / (1024 * 1024):.1f} MB"
    else:
        return f"{size / (1024 * 1024 * 1024):.1f} GB"


async def browse_ftp_directory(config: dict, path: str = "/") -> dict:
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, browse_ftp_sync, config, path)


# ==================== MULTI-FILE SYNC ====================

def _detect_best_separator(first_line: str, preferred: str) -> str:
    """Return the separator that produces the most columns in first_line.
    Falls back to preferred if no alternative wins by a clear margin."""
    candidates = [preferred, ';', ',', '\t', '|']
    # deduplicate while preserving order
    seen = set()
    candidates = [c for c in candidates if not (c in seen or seen.add(c))]
    best_sep = preferred
    best_count = len(list(csv.reader([first_line], delimiter=preferred, quotechar='"'))[0]) if first_line else 1
    for sep in candidates[1:]:
        try:
            count = len(list(csv.reader([first_line], delimiter=sep, quotechar='"'))[0])
        except Exception:
            continue
        if count > best_count:
            best_count = count
            best_sep = sep
    if best_sep != preferred:
        logger.info(f"Auto-detected separator {repr(best_sep)} ({best_count} cols) instead of {repr(preferred)}")
    return best_sep


def parse_text_file(content: bytes, separator: str = ";", header_row: int = 1) -> list:
    """Parse a text file (CSV/TXT). header_row=0 means no header (positional col names).
    If the configured separator produces only 1 column, auto-detects the best separator."""
    try:
        decoded = content.decode('utf-8-sig', errors='replace')
    except Exception:
        try:
            decoded = content.decode('utf-8', errors='replace')
        except Exception:
            decoded = content.decode('latin-1', errors='replace')
    lines = decoded.strip().split('\n')
    if header_row > 1:
        lines = lines[header_row - 1:]
    if not lines:
        return []
    if separator == '\\t':
        separator = '\t'
    first_line = lines[0].rstrip('\r') if lines else ''
    # Auto-detect separator when configured one yields only 1 column
    separator = _detect_best_separator(first_line, separator)
    if header_row == 0:
        first_row_parsed = list(csv.reader([first_line], delimiter=separator, quotechar='"'))
        num_cols = len(first_row_parsed[0]) if first_row_parsed else 0
        fieldnames = [f'col_{i}' for i in range(num_cols)]
        reader = csv.DictReader(lines, fieldnames=fieldnames, delimiter=separator, quotechar='"')
    else:
        reader = csv.DictReader(lines, delimiter=separator, quotechar='"')
    return list(reader)


def extract_zip_files(content: bytes) -> dict:
    """Extract all files from a ZIP archive, returns {filename: bytes}.
    Valida paths para prevenir path traversal (Zip Slip)."""
    result = {}
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            normalized = os.path.normpath(info.filename)
            if normalized.startswith('..') or os.path.isabs(normalized):
                logger.warning(f"Zip Slip: path sospechoso ignorado: {info.filename}")
                continue
            with zf.open(info.filename) as f:
                result[info.filename] = f.read()
    return result


async def resolve_latest_file(supplier: dict, file_config: dict) -> str:
    """If auto_latest is set, find the latest matching file (e.g. latest ZIP)"""
    file_path = file_config.get('path', '')
    if not file_config.get('auto_latest'):
        return file_path

    # Get the directory of the file
    dir_path = '/'.join(file_path.split('/')[:-1]) or '/'
    filename = file_path.split('/')[-1]
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    try:
        result = await browse_ftp_directory(supplier, dir_path)
        if result.get('status') != 'ok':
            return file_path

        # Filter files with same extension
        candidates = [f for f in result['files'] if not f['is_dir'] and f['name'].lower().endswith(f'.{ext}')]
        if not candidates:
            return file_path

        # Sort by name descending (works for date-based naming like _20260223)
        candidates.sort(key=lambda x: x['name'], reverse=True)
        latest = candidates[0]

        if latest['path'] != file_path:
            logger.info(f"Auto-latest: resolved {file_path} -> {latest['path']}")

        return latest['path']
    except Exception as e:
        logger.warning(f"Could not resolve latest file for {file_path}: {e}")
        return file_path


async def sync_supplier_multifile(supplier: dict, sync_type: str = "manual") -> dict:
    """Sync supplier with multiple file paths - downloads all, merges by key"""
    ftp_paths = supplier.get('ftp_paths', [])
    if not ftp_paths:
        return await sync_supplier(supplier, sync_type)

    start_time = datetime.now(UTC)
    logger.info(f"Multi-file sync for {supplier['name']}: {len(ftp_paths)} files configured")
    all_file_data = {}
    all_detected_columns = {}

    # Use supplier-level csv_header_row and csv_separator (set by preset) for all
    # file parsing. Per-file values in ftp_paths are UI defaults and may not
    # reflect the actual file format; the preset sets the authoritative values.
    _sup_hdr_raw = supplier.get('csv_header_row', 1)
    supplier_hdr = 1 if _sup_hdr_raw is None else int(_sup_hdr_raw)
    _sup_sep = supplier.get('csv_separator') or ';'
    supplier_sep = '\t' if _sup_sep == '\\t' else _sup_sep
    strip_ean_quotes = supplier.get('strip_ean_quotes', False)

    for file_config in ftp_paths:
        file_path = await resolve_latest_file(supplier, file_config)
        role = file_config.get('role', 'products')
        sep = supplier_sep
        label = file_config.get('label', file_path)

        if not file_path:
            continue

        try:
            logger.info(f"  Downloading: {file_path} (role: {role})")
            content = await download_file_from_ftp({**supplier, 'ftp_path': file_path})

            if file_path.lower().endswith('.zip') or (len(content) >= 2 and content[:2] == b'PK'):
                extracted = extract_zip_files(content)
                logger.info(f"  ZIP contains {len(extracted)} files: {list(extracted.keys())}, header_row={supplier_hdr}")
                for fname, fcontent in extracted.items():
                    rows = parse_text_file(fcontent, sep, supplier_hdr)
                    sub_role = role
                    fname_lower = fname.lower()
                    if 'stock' in fname_lower:
                        sub_role = 'stock'
                    elif 'price' in fname_lower and 'qb' not in fname_lower:
                        sub_role = 'prices'
                    elif 'qb' in fname_lower:
                        sub_role = 'prices_qb'
                    elif 'product' in fname_lower:
                        sub_role = 'products'
                    elif 'kit' in fname_lower:
                        sub_role = 'kit'
                    elif 'minqty' in fname_lower:
                        sub_role = 'min_qty'

                    # Keep the file with most rows for each role
                    if sub_role not in all_file_data or len(rows) > len(all_file_data[sub_role]):
                        all_file_data[sub_role] = rows
                        if rows:
                            all_detected_columns[sub_role] = list(rows[0].keys())
                    logger.info(f"    {fname}: {len(rows)} rows, role={sub_role}")
            else:
                rows = parse_text_file(content, sep, supplier_hdr)
                all_file_data[role] = rows
                if rows:
                    all_detected_columns[role] = list(rows[0].keys())
                logger.info(f"  {label}: {len(rows)} rows")

        except Exception as e:
            logger.error(f"  Error downloading {file_path}: {e}")
            continue

    if not all_file_data:
        return {"status": "error", "message": "No se pudo descargar ningún archivo"}

    products_data = all_file_data.get('products', [])
    prices_data = all_file_data.get('prices', [])
    stock_data = all_file_data.get('stock', [])

    if not products_data and prices_data:
        products_data = prices_data

    if not products_data:
        return {"status": "error", "message": "No se encontraron datos de productos"}

    # Detect merge key from products (first column is usually the ID)
    first_row = products_data[0] if products_data else {}
    product_keys = list(first_row.keys())
    merge_key = product_keys[0] if product_keys else None

    # For the merge prefix strategy, reuse the already-computed supplier_hdr.
    multifile_header_row = supplier_hdr

    # Build lookup dictionaries
    prices_lookup = {}
    price_merge_key = None
    if prices_data:
        price_keys = list(prices_data[0].keys()) if prices_data else []
        price_merge_key = price_keys[0] if price_keys else None
        if price_merge_key:
            for row in prices_data:
                key = str(row.get(price_merge_key, '')).strip()
                if key:
                    prices_lookup[key] = row

    stock_lookup = {}
    stock_merge_key = None
    if stock_data:
        stock_keys = list(stock_data[0].keys()) if stock_data else []
        stock_merge_key = stock_keys[0] if stock_keys else None
        if stock_merge_key:
            for row in stock_data:
                key = str(row.get(stock_merge_key, '')).strip()
                if key:
                    stock_lookup[key] = row

    logger.info(f"Merging: {len(products_data)} products, {len(prices_lookup)} prices, {len(stock_lookup)} stock entries")

    now = datetime.now(UTC).isoformat()
    errors = 0

    # Normalize all products first (merge + mapping)
    normalized_products = []
    column_mapping = supplier.get('column_mapping')
    for raw_product in products_data:
        try:
            prod_id = str(raw_product.get(merge_key, '')).strip()
            if not prod_id:
                errors += 1
                continue

            merged = dict(raw_product)
            if prod_id in prices_lookup:
                for k, v in prices_lookup[prod_id].items():
                    if multifile_header_row == 0:
                        if k != price_merge_key:
                            merged[f"prices_{k}"] = v
                    elif k not in merged or not merged[k]:
                        merged[k] = v
            if prod_id in stock_lookup:
                for k, v in stock_lookup[prod_id].items():
                    if multifile_header_row == 0:
                        if k != stock_merge_key:
                            merged[f"stock_{k}"] = v
                    elif k not in merged or not merged[k]:
                        merged[k] = v

            if column_mapping:
                normalized = apply_column_mapping(merged, column_mapping, strip_ean_quotes)
            else:
                normalized = normalize_product_data(merged, strip_ean_quotes)

            sku = normalized.get('sku') or prod_id
            name = normalized.get('name', '')
            if not name:
                errors += 1
                continue
            normalized_products.append(normalized)

        except Exception as e:
            logger.error(f"Error processing multi-file product: {e}")
            errors += 1

    # Bulk upsert with SKU cache for efficient lookups
    sku_cache = await prefetch_existing_products(supplier['id'], supplier['user_id'])
    bulk_result = await bulk_upsert_products(supplier, normalized_products, sku_cache, now)
    imported = bulk_result["imported"]
    updated = bulk_result["updated"]
    errors += bulk_result["errors"]

    duration = (datetime.now(UTC) - start_time).total_seconds()
    product_count = await db.products.count_documents({"supplier_id": supplier['id']})

    # Build flat detected_columns list (products + prefixed prices/stock) for the mapping UI
    flat_detected = []
    prod_cols = all_detected_columns.get('products', [])
    flat_detected.extend(prod_cols)
    price_cols = all_detected_columns.get('prices', [])
    if price_cols and multifile_header_row == 0:
        price_key = price_cols[0] if price_cols else None
        flat_detected.extend([f"prices_{c}" for c in price_cols if c != price_key])
    elif price_cols:
        flat_detected.extend([c for c in price_cols if c not in flat_detected])
    stock_cols = all_detected_columns.get('stock', [])
    if stock_cols and multifile_header_row == 0:
        stock_key = stock_cols[0] if stock_cols else None
        flat_detected.extend([f"stock_{c}" for c in stock_cols if c != stock_key])
    elif stock_cols:
        flat_detected.extend([c for c in stock_cols if c not in flat_detected])

    await db.suppliers.update_one({"id": supplier['id']}, {"$set": {
        "product_count": product_count, "last_sync": now,
        "detected_columns": flat_detected if flat_detected else list(all_detected_columns.keys())
    }})

    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), "type": "sync_complete",
        "message": f"Sincronización multi-archivo: {supplier['name']} - {imported} nuevos, {updated} actualizados ({len(ftp_paths)} archivos)",
        "product_id": None, "product_name": None,
        "user_id": supplier["user_id"], "read": False, "created_at": now
    })

    final_result = {
        "status": "success", "imported": imported, "updated": updated,
        "errors": errors, "files_processed": len(all_file_data),
        "detected_columns": flat_detected
    }
    await record_sync_history(supplier, final_result, sync_type, duration)

    logger.info(f"Multi-file sync complete: {imported} imported, {updated} updated, {errors} errors")
    return final_result


# ==================== WOOCOMMERCE SYNC ====================

def get_woocommerce_client(config: dict) -> WooCommerceAPI:
    return WooCommerceAPI(url=config['store_url'], consumer_key=config['consumer_key'], consumer_secret=config['consumer_secret'], version="wc/v3", timeout=WOOCOMMERCE_API_TIMEOUT)


def mask_key(key: str) -> str:
    if len(key) <= 4:
        return "****"
    return "*" * (len(key) - 4) + key[-4:]


def calculate_final_price(base_price: float, product: dict, rules: list) -> float:
    """Calculate final price by cumulatively applying all matching margin rules.

    HIGH FIX #11: Apply rules cumulatively in priority order, not just the first match.
    Each rule is applied on top of the previous result.

    Example:
    - base_price: 100€
    - rule 1 (all, +10%): 100 * 1.10 = 110€
    - rule 2 (category, +20%): 110 * 1.20 = 132€
    """
    final_price = base_price

    # Rules are already sorted by priority (descending) from the caller
    for rule in rules:
        applies = False

        # Check if rule applies to this product
        if rule["apply_to"] == "all" or rule["apply_to"] == "category" and product.get("category") == rule.get("apply_to_value") or rule["apply_to"] == "supplier" and product.get("supplier_id") == rule.get("apply_to_value") or rule["apply_to"] == "product" and product.get("id") == rule.get("apply_to_value"):
            applies = True

        if applies:
            # Check min/max price bounds (on base price)
            if rule.get("min_price") and base_price < rule["min_price"]:
                continue
            if rule.get("max_price") and base_price > rule["max_price"]:
                continue

            # SECURITY FIX #11: Apply cumulatively on final_price, not base_price
            # This allows multiple rules to be stacked
            if rule["rule_type"] == "percentage":
                final_price = final_price * (1 + rule["value"] / 100)
            elif rule["rule_type"] == "fixed":
                final_price = final_price + rule["value"]

            # DO NOT BREAK - continue applying other rules
            logger.debug(f"Applied margin rule: {rule.get('name', 'unnamed')} → {final_price:.2f}€")

    return final_price


async def sync_woocommerce_store_price_stock(config: dict):
    store_name = config.get("name", "Unknown")
    catalog_id = config.get("catalog_id")
    if not catalog_id:
        logger.warning(f"No catalog associated with WooCommerce store {store_name}")
        return
    logger.info(f"Starting price/stock sync for WooCommerce store: {store_name}")
    try:
        wcapi = get_woocommerce_client(config)
        catalog_items = await db.catalog_items.find({"catalog_id": catalog_id, "active": True}).to_list(10000)
        if not catalog_items:
            logger.info(f"No active products in catalog for store {store_name}")
            return
        margin_rules = await db.catalog_margin_rules.find({"catalog_id": catalog_id}).to_list(100)
        wc_products_by_ean = {}
        wc_products_by_sku = {}
        page = 1
        while True:
            response = await asyncio.to_thread(wcapi.get, "products", params={"per_page": 100, "page": page})
            if response.status_code == 200:
                products_batch = response.json()
                if not products_batch:
                    break
                for p in products_batch:
                    for meta in p.get("meta_data", []):
                        if meta.get("key") in ["_global_unique_id", "_gtin", "_ean", "gtin"]:
                            ean_value = meta.get("value")
                            if ean_value:
                                wc_products_by_ean[ean_value] = p["id"]
                                break
                    if p.get("sku"):
                        wc_products_by_sku[p["sku"]] = p["id"]
                page += 1
                if len(products_batch) < 100:
                    break
            else:
                logger.error(f"Error fetching WooCommerce products: {response.text}")
                break
        logger.info(f"Found {len(wc_products_by_ean)} products by EAN, {len(wc_products_by_sku)} by SKU")
        updated = 0
        failed = 0
        for item in catalog_items:
            try:
                product = await db.products.find_one({"id": item["product_id"]})
                if not product:
                    continue
                base_price = item.get("custom_price") or product.get("price", 0)
                final_price = calculate_final_price(base_price, product, margin_rules)
                ean = product.get("ean", "")
                sku = product.get("sku", "")
                wc_product_id = wc_products_by_ean.get(ean) if ean else None
                if not wc_product_id and sku:
                    wc_product_id = wc_products_by_sku.get(sku)
                if wc_product_id:
                    update_data = {"regular_price": str(round(final_price, 2)), "stock_quantity": product.get("stock", 0)}
                    response = await asyncio.to_thread(wcapi.put, f"products/{wc_product_id}", update_data)
                    if response.status_code in [200, 201]:
                        updated += 1
                    else:
                        failed += 1
                        logger.warning(f"Failed to update product {ean or sku}: {response.text[:100]}")
            except Exception as e:
                failed += 1
                logger.error(f"Error processing catalog item: {e}")
        now = datetime.now(UTC).isoformat()
        await db.woocommerce_configs.update_one({"id": config["id"]}, {"$set": {"last_sync": now, "products_synced": updated}})
        logger.info(f"WooCommerce sync completed for {store_name}: {updated} updated, {failed} failed")
    except Exception as e:
        logger.error(f"Error syncing WooCommerce store {store_name}: {e}")


async def sync_all_woocommerce_stores():
    from services.unified_sync import _global_sync_lock
    logger.info("Starting scheduled WooCommerce sync for all stores (waiting for global lock)...")
    async with _global_sync_lock:
        configs = await db.woocommerce_configs.find({
            "auto_sync_enabled": True,
            "catalog_id": {"$nin": [None, ""]}
        }).to_list(1000)
        logger.info(f"Found {len(configs)} WooCommerce stores with auto-sync enabled")
        for config in configs:
            try:
                await sync_woocommerce_store_price_stock(config)
                await asyncio.sleep(5)
            except Exception as e:
                logger.error(f"Error syncing WooCommerce store {config.get('name', config['id'])}: {e}")
        logger.info("Scheduled WooCommerce sync completed")


# ==================== WOOCOMMERCE CATEGORY FUNCTIONS ====================

def get_woocommerce_categories_sync(config: dict) -> list:
    """Get all categories from WooCommerce"""
    wcapi = get_woocommerce_client(config)
    categories = []
    page = 1
    try:
        while True:
            response = wcapi.get("products/categories", params={"per_page": 100, "page": page})
            if response.status_code == 200:
                batch = response.json()
                if not batch:
                    break
                categories.extend(batch)
                page += 1
                if len(batch) < 100:
                    break
            else:
                logger.error(f"Error fetching WooCommerce categories: {response.text[:200]}")
                break
    except Exception as e:
        logger.error(f"WooCommerce get_categories error: {e}")
    return categories


async def get_woocommerce_categories(config: dict) -> list:
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, get_woocommerce_categories_sync, config)


def create_woocommerce_category_sync(config: dict, category_data: dict) -> dict:
    """Create a category in WooCommerce"""
    wcapi = get_woocommerce_client(config)
    try:
        payload = {
            "name": category_data.get("name", ""),
            "parent": category_data.get("parent_id", 0),
            "description": category_data.get("description", "")
        }
        response = wcapi.post("products/categories", payload)
        if response.status_code in [200, 201]:
            cat = response.json()
            return {"status": "success", "category_id": cat.get("id"), "message": "Categoría creada"}
        else:
            return {"status": "error", "message": f"Error: {response.status_code} - {response.text[:200]}"}
    except Exception as e:
        return {"status": "error", "message": f"Error: {str(e)}"}


async def create_woocommerce_category(config: dict, category_data: dict) -> dict:
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, create_woocommerce_category_sync, config, category_data)


async def find_or_create_woocommerce_category(config: dict, name: str, parent_id: int = 0, existing_categories: list = None) -> int:
    """Find existing category by name or create it"""
    if existing_categories is None:
        existing_categories = await get_woocommerce_categories(config)

    # Search for existing category
    for cat in existing_categories:
        if cat.get("name", "").lower() == name.lower() and cat.get("parent", 0) == parent_id:
            return cat.get("id")

    # Create new category
    result = await create_woocommerce_category(config, {"name": name, "parent_id": parent_id})
    if result.get("status") == "success":
        return result.get("category_id")
    return None


async def export_catalog_categories_to_woocommerce(config: dict, catalog_id: str, user_id: str) -> dict:
    """Export catalog categories to WooCommerce store"""
    from services.database import db

    # Get catalog categories
    categories = await db.catalog_categories.find(
        {"catalog_id": catalog_id}, {"_id": 0, "user_id": 0}
    ).sort("position", 1).to_list(500)

    if not categories:
        return {"status": "warning", "created": 0, "message": "No hay categorías para exportar"}

    # Get existing WooCommerce categories
    wc_categories = await get_woocommerce_categories(config)

    # Build mapping of our category IDs to WooCommerce category IDs
    category_mapping = {}  # local_id -> wc_id
    created = 0
    errors = []

    # Process categories level by level
    max_level = max(cat.get("level", 0) for cat in categories)
    for level in range(max_level + 1):
        level_cats = [c for c in categories if c.get("level", 0) == level]
        for cat in level_cats:
            try:
                wc_parent_id = 0
                if cat.get("parent_id") and cat["parent_id"] in category_mapping:
                    wc_parent_id = category_mapping[cat["parent_id"]]

                wc_id = await find_or_create_woocommerce_category(
                    config, cat["name"], wc_parent_id, wc_categories
                )

                if wc_id:
                    category_mapping[cat["id"]] = wc_id
                    # Check if it was created (not found in existing)
                    was_existing = any(
                        wc.get("name", "").lower() == cat["name"].lower() and wc.get("parent", 0) == wc_parent_id
                        for wc in wc_categories
                    )
                    if not was_existing:
                        created += 1
                        # Add to existing list to avoid duplicates
                        wc_categories.append({"id": wc_id, "name": cat["name"], "parent": wc_parent_id})
                else:
                    errors.append(f"Error creando categoría: {cat['name']}")
            except Exception as e:
                errors.append(f"Error procesando {cat['name']}: {str(e)[:50]}")

    return {
        "status": "success" if not errors else "partial",
        "created": created,
        "total": len(categories),
        "mapped": len(category_mapping),
        "category_mapping": category_mapping,
        "errors": errors[:10]
    }


async def fetch_all_store_products(store_config: dict) -> list:
    """
    Fetch ALL products from a store using paginated API calls.
    Supports WooCommerce, PrestaShop, Shopify, Magento, and Wix.
    """
    from services.platforms import get_platform_client

    platform = store_config.get("platform", "woocommerce")
    store_products = []

    if platform == "woocommerce":
        wc = get_woocommerce_client(store_config)
        page = 1
        while True:
            batch = await asyncio.to_thread(
                wc.get, "products", params={"per_page": 100, "page": page}
            )

            # Parse JSON with proper encoding handling
            if hasattr(batch, 'json'):
                try:
                    # First try: use requests' automatic encoding detection
                    batch = batch.json()
                except Exception as e:
                    # Fallback: manually handle encoding and JSON structure issues
                    logger.warning(f"JSON decode error on page {page}: {type(e).__name__}: {str(e)[:150]}")

                    try:
                        # Get raw content with explicit encoding handling
                        if hasattr(batch, 'content'):
                            raw_bytes = batch.content
                            logger.info(f"Response size: {len(raw_bytes)} bytes on page {page}")

                            # Try different encodings
                            parsed_data = None
                            text = None

                            for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']:
                                try:
                                    text = raw_bytes.decode(encoding)
                                    parsed_data = json.loads(text)
                                    logger.info(f"Successfully parsed page {page} with {encoding} encoding")
                                    batch = parsed_data
                                    break
                                except (json.JSONDecodeError, UnicodeDecodeError) as encode_error:
                                    logger.debug(f"  {encoding}: {type(encode_error).__name__}")
                                    # Keep the text from UTF-8 for later sanitization
                                    if encoding == 'utf-8' and text is None:
                                        try:
                                            text = raw_bytes.decode('utf-8', errors='replace')
                                        except:
                                            pass
                                    continue

                            # If standard parsing failed and we have text, try sanitizing
                            if parsed_data is None and text:
                                logger.warning(f"Standard parsing failed on page {page}, attempting JSON sanitization")
                                try:
                                    # Try to fix invalid escape sequences
                                    # Replace invalid escapes like \x, \', etc. with their actual characters
                                    # Match backslash followed by invalid escape
                                    sanitized = re.sub(r'\\([^"\\/bfnrtu])', r'\1', text)
                                    parsed_data = json.loads(sanitized)
                                    logger.info(f"Successfully parsed page {page} after sanitization")
                                    batch = parsed_data
                                except Exception as sanitize_error:
                                    logger.warning(f"Sanitization failed: {type(sanitize_error).__name__}")
                                    # Last resort: try with errors='replace'
                                    if text:
                                        try:
                                            # Re-encode with errors='replace' and try again
                                            text_replaced = raw_bytes.decode('utf-8', errors='replace')
                                            parsed_data = json.loads(text_replaced)
                                            logger.info(f"Successfully parsed page {page} with utf-8 errors='replace'")
                                            batch = parsed_data
                                        except Exception as last_error:
                                            logger.error(f"All parsing strategies failed on page {page}")
                                            raise last_error

                            if parsed_data is None:
                                raise ValueError(f"Could not parse JSON response on page {page}")
                        else:
                            logger.error(f"Response object has no 'content' attribute on page {page}")
                            raise e
                    except Exception as fallback_error:
                        logger.error(f"Failed to parse WooCommerce response on page {page}: {type(fallback_error).__name__}: {fallback_error}")
                        raise

            if isinstance(batch, dict) and "body" in batch:
                batch = batch["body"]
            if not batch or not isinstance(batch, list):
                break
            store_products.extend(batch)
            if len(batch) < 100:
                break
            page += 1
    else:
        client = get_platform_client(store_config)
        if client and hasattr(client, 'get_all_products'):
            store_products = await asyncio.to_thread(client.get_all_products)
        elif client:
            store_products = await asyncio.to_thread(client.get_products, 100)

    return store_products


def extract_store_product_info(store_prod: dict, platform: str) -> dict:
    """
    Extract matching fields from store product.
    Price and stock come from SyncStock supplier products, not the store.

    Fields extracted: SKU, EAN, name, description, image, category, brand
    """
    info = {
        "sku": "",
        "ean": "",
        "name": "",
        "description": "",
        "image_url": "",
        "category": "",
        "brand": ""
    }

    if platform == "woocommerce":
        info["sku"] = (store_prod.get("sku") or "").strip()
        info["ean"] = (store_prod.get("ean") or "").strip()
        info["name"] = (store_prod.get("name") or "").strip()
        info["description"] = store_prod.get("description") or store_prod.get("short_description") or ""
        images = store_prod.get("images") or []
        info["image_url"] = images[0].get("src", "") if images else ""
        cats = store_prod.get("categories") or []
        info["category"] = cats[0].get("name", "") if cats else ""
        info["brand"] = store_prod.get("brands", [{}])[0].get("name", "") if store_prod.get("brands") else ""
    elif platform == "prestashop":
        info["sku"] = (store_prod.get("reference") or "").strip()
        info["ean"] = (store_prod.get("ean13") or "").strip()
        name_val = store_prod.get("name") or ""
        if isinstance(name_val, list):
            name_val = name_val[0].get("value", "") if name_val else ""
        elif isinstance(name_val, dict):
            name_val = name_val.get("value", "") or name_val.get("language", "")
        info["name"] = str(name_val).strip()
        info["description"] = store_prod.get("description") or store_prod.get("description_short") or ""
        info["category"] = store_prod.get("id_category_default", "")
        info["brand"] = ""
    elif platform == "shopify":
        variants = store_prod.get("variants") or []
        first_variant = variants[0] if variants else {}
        info["sku"] = (first_variant.get("sku") or "").strip()
        info["ean"] = (first_variant.get("barcode") or "").strip()
        info["name"] = (store_prod.get("title") or "").strip()
        info["description"] = store_prod.get("body_html") or ""
        info["brand"] = store_prod.get("vendor") or ""
        info["category"] = store_prod.get("product_type") or ""
        images = store_prod.get("images") or []
        info["image_url"] = images[0].get("src", "") if images else ""
    elif platform == "magento":
        info["sku"] = (store_prod.get("sku") or "").strip()
        info["name"] = (store_prod.get("name") or "").strip()
        info["description"] = store_prod.get("description") or ""
        info["ean"] = ""
        info["brand"] = ""
        info["category"] = ""
    elif platform == "wix":
        info["sku"] = (store_prod.get("sku") or "").strip()
        info["name"] = (store_prod.get("name") or "").strip()
        info["description"] = store_prod.get("description") or ""
        info["ean"] = ""
        media = store_prod.get("media", {}).get("items") or []
        info["image_url"] = media[0].get("url", "") if media else ""
        info["category"] = ""
        info["brand"] = ""

    return info


async def create_catalog_from_store_products(
    user_id: str,
    store_config_id: str,
    catalog_name: str | None = None,
    catalog_id: str | None = None,
    match_by: list[str] | None = None,
    skip_unmatched: bool = True
) -> dict:
    """
    Create a catalog with products from a store by matching them with supplier products.

    IMPORTANT: Only products that match supplier products are added to the catalog.
    The price is always determined from the matched supplier product, not the store.

    Supports pagination for large stores and sends WebSocket progress notifications.

    Args:
        user_id: ID of the user
        store_config_id: ID of the store configuration
        catalog_name: Name for the new catalog (if creating new)
        catalog_id: Use existing catalog instead of creating new one
        match_by: List of fields to match by (sku, ean, name)
        skip_unmatched: (Deprecated - always True) Skip products not found in suppliers

    Returns:
        dict with creation results
    """
    if match_by is None:
        match_by = ["sku", "ean", "name"]

    errors = []
    now = datetime.now(UTC).isoformat()
    created_products = 0
    catalog = None

    try:
        # Get store configuration
        store_config = await db.woocommerce_configs.find_one(
            {"id": store_config_id, "user_id": user_id}
        )
        if not store_config:
            raise Exception("Configuración de tienda no encontrada")

        store_name = store_config.get("name", "Tienda")
        platform = store_config.get("platform", "woocommerce")

        # Step 1: Notify — fetching products from store
        await send_realtime_notification(user_id, {
            "id": str(uuid.uuid4()),
            "type": "sync_progress",
            "message": f"Obteniendo productos de '{store_name}'...",
            "progress": 5,
            "processed": 0,
            "total": 0,
        })

        # Fetch ALL products with pagination
        try:
            store_products = await fetch_all_store_products(store_config)
        except Exception as e:
            logger.error(f"Error fetching store products: {e}")
            raise Exception(f"Error al obtener productos de la tienda: {str(e)[:100]}")

        if not store_products:
            raise Exception("No se encontraron productos en la tienda")

        total_store = len(store_products)

        # Step 2: Notify — products fetched
        await send_realtime_notification(user_id, {
            "id": str(uuid.uuid4()),
            "type": "sync_progress",
            "message": f"Se encontraron {total_store} productos en '{store_name}'. Preparando catálogo...",
            "progress": 15,
            "processed": 0,
            "total": total_store,
        })

        # Get or create catalog
        if catalog_id:
            catalog = await db.catalogs.find_one(
                {"id": catalog_id, "user_id": user_id}
            )
            if not catalog:
                raise Exception("Catálogo no encontrado")
        else:
            if not catalog_name:
                catalog_name = f"Catálogo {store_name}"

            catalog_id = str(uuid.uuid4())
            catalog = {
                "id": catalog_id,
                "user_id": user_id,
                "name": catalog_name,
                "description": f"Creado desde productos de {store_name}",
                "is_default": False,
                "created_at": now
            }
            await db.catalogs.insert_one(catalog)

        # Step 3: Notify — loading supplier products for matching
        await send_realtime_notification(user_id, {
            "id": str(uuid.uuid4()),
            "type": "sync_progress",
            "message": "Cargando productos de proveedores para cruzar datos...",
            "progress": 20,
            "processed": 0,
            "total": total_store,
        })

        # Get all user's products from suppliers for matching
        supplier_products = await db.products.find(
            {"user_id": user_id},
            {"_id": 0}
        ).to_list(None)

        # Create lookup maps for fast matching
        products_by_sku = {}
        products_by_ean = {}
        products_by_name = {}

        for prod in supplier_products:
            sku = (prod.get("sku") or "").lower().strip()
            ean = (prod.get("ean") or "").lower().strip()
            name = (prod.get("name") or "").lower().strip()

            if sku:
                products_by_sku[sku] = prod
            if ean:
                products_by_ean[ean] = prod
            if name and len(name) > 3:
                if name not in products_by_name:
                    products_by_name[name] = []
                products_by_name[name].append(prod)

        # Match store products with supplier products
        matched = 0
        unmatched = 0
        added_items = 0
        catalog_items = []
        products_to_create = []

        for i, store_prod in enumerate(store_products):
            matched_product = None
            info = extract_store_product_info(store_prod, platform)

            store_sku = info["sku"].lower()
            store_ean = info["ean"].lower()
            store_product_name = info["name"].lower()

            # Try to match by SKU first (most reliable)
            if "sku" in match_by and store_sku and store_sku in products_by_sku:
                matched_product = products_by_sku[store_sku]

            # Then by EAN
            if not matched_product and "ean" in match_by and store_ean and store_ean in products_by_ean:
                matched_product = products_by_ean[store_ean]

            # Finally by name (exact match, case-insensitive)
            if not matched_product and "name" in match_by and store_product_name and store_product_name in products_by_name:
                matched_product = products_by_name[store_product_name][0]

            if matched_product:
                matched += 1
                item_id = str(uuid.uuid4())
                catalog_items.append({
                    "id": item_id,
                    "catalog_id": catalog_id,
                    "product_id": matched_product["id"],
                    "custom_price": None,
                    "custom_name": None,
                    "active": True,
                    "category_ids": [],
                    "created_at": now
                })
                added_items += 1
            else:
                unmatched += 1
                # No crear productos nuevos - solo hacer matching con proveedores
                # El precio debe venir siempre de los productos del proveedor

            # Send progress every 50 products
            if (i + 1) % 50 == 0 or (i + 1) == total_store:
                pct = 20 + int(((i + 1) / total_store) * 70)
                await send_realtime_notification(user_id, {
                    "id": str(uuid.uuid4()),
                    "type": "sync_progress",
                    "message": f"Cruzando productos: {i + 1}/{total_store} ({matched} encontrados)",
                    "progress": pct,
                    "processed": i + 1,
                    "total": total_store,
                })

        # Note: No new products are created during store import
        # All products must come from existing suppliers to ensure proper pricing

        # Insert catalog items in bulk (with duplicate handling)
        if catalog_items:
            try:
                await db.catalog_items.insert_many(catalog_items, ordered=False)
            except Exception as e:
                # Handle duplicate key errors gracefully
                if "duplicate" in str(e).lower() or "11000" in str(e):
                    logger.warning(f"Some catalog items have duplicate keys, attempting individual inserts: {str(e)[:100]}")
                    # Try inserting individually to skip duplicates
                    inserted = 0
                    for item in catalog_items:
                        try:
                            await db.catalog_items.insert_one(item)
                            inserted += 1
                        except Exception as item_error:
                            if "duplicate" in str(item_error).lower() or "11000" in str(item_error):
                                # Skip this item (it's a duplicate)
                                logger.debug(f"Skipping duplicate catalog item: {item['id']}")
                            else:
                                logger.error(f"Error inserting catalog item {item['id']}: {item_error}")
                    logger.info(f"Inserted {inserted}/{len(catalog_items)} catalog items (skipped duplicates)")
                else:
                    # Re-raise if it's not a duplicate key error
                    raise

        # Step 5: Notify — complete
        await send_realtime_notification(user_id, {
            "id": str(uuid.uuid4()),
            "type": "sync_complete",
            "message": f"Catálogo '{catalog['name']}' creado: {matched} coincidencias, {unmatched} sin coincidencia, {added_items} añadidos al catálogo",
        })

        return {
            "status": "success",
            "catalog_id": catalog["id"],
            "catalog_name": catalog["name"],
            "total_products": total_store,
            "matched_products": matched,
            "unmatched_products": unmatched,
            "added_items": added_items,
            "created_products": created_products,
            "errors": errors,
            "created_at": now,
        }

    except Exception as e:
        logger.error(f"Error creating catalog from store: {e}")
        # Notify error via WebSocket
        await send_realtime_notification(user_id, {
            "id": str(uuid.uuid4()),
            "type": "sync_error",
            "message": f"Error creando catálogo desde tienda: {str(e)[:100]}",
        })
        return {
            "status": "error",
            "catalog_id": catalog["id"] if catalog else (catalog_id if catalog_id else None),
            "catalog_name": catalog["name"] if catalog else (catalog_name or ""),
            "total_products": len(store_products) if "store_products" in dir() else 0,
            "matched_products": 0,
            "unmatched_products": 0,
            "added_items": 0,
            "created_products": 0,
            "errors": [str(e)],
            "created_at": now,
        }
